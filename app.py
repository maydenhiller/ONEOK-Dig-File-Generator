"""ONEOK Dig File Generator - single file Streamlit app.

Upload a staking report template, a cheat sheet template, one or more dig
sheets, and optionally the alignment sheets and a pipeline KMZ. Get back a
filled cheat sheet plus one staking report per dig, named
"<Dig Name>_Staking Report.xlsm".

Everything the uploads can establish is filled in. Everything they cannot is
left as Unknown or blank and flagged in the review table - a wrong value in a
signed report is worse than a missing one.

Deliberately kept as one file with no local package imports: Streamlit Cloud
deployments break silently when a subpackage does not make it into the repo.

Latitude, longitude, elevation, EDOC, survey date and photos are field
measurements and stay manual.
"""

from __future__ import annotations

import datetime as _dt
import io
import math
import re
import sys
import zipfile
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from dataclasses import dataclass, field
from datetime import date
from typing import Iterable, Optional
from xml.etree import ElementTree
from xml.sax.saxutils import escape

import requests
import streamlit as st
from PIL import Image, ImageDraw, ImageFont


# ==========================================================================
# MODELS
# ==========================================================================
# Shared data structures for the ONEOK dig file generator.

UNKNOWN = "Unknown"


def station_to_feet(value) -> Optional[int]:
    """Normalise a station to whole feet.

    Accepts the two forms that show up across dig sheets:
      * ``3401+57``  (PDF dig sheets)  -> 340157
      * ``243742``   (xlsx dig sheets) -> 243742
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if "+" in text:
        left, _, right = text.partition("+")
        try:
            return int(left) * 100 + int(round(float(right)))
        except ValueError:
            return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def feet_to_station(feet) -> str:
    """340157 -> '3401+57'. Used for matching against alignment sheet ranges."""
    if feet is None:
        return ""
    feet = int(round(float(feet)))
    return f"{feet // 100}+{feet % 100:02d}"


AGM_NAME_RE = re.compile(r"^(AGM\s*\d+)", re.I)
VALVE_NAME_RE = re.compile(
    r"^((?:LAUNCH|LAUNCHER|RECEIVE|RECEIVER|MAINLINE|MAIN\s*LINE|CHECK|BLOCK)\s+VALVE)",
    re.I,
)


def agm_name(reference) -> Optional[str]:
    """The short reference name that goes on the report.

    The cell holds a full description - "AGM 520, Sta. 2635+78, 20' N of C/L
    NE 80 Rd." or "LAUNCH VALVE Danville, Sta. 2582+14" - and the report wants
    just "AGM 520" or "Launch Valve", so the reference type is taken off the
    front rather than everything up to the first comma, which would keep the
    location too.
    """
    if reference is None:
        return None
    text = re.sub(r"\s+", " ", str(reference)).strip()
    if not text:
        return None

    match = AGM_NAME_RE.match(text)
    if match:
        return re.sub(r"\s+", " ", match.group(1)).upper()

    match = VALVE_NAME_RE.match(text)
    if match:
        return re.sub(r"\s+", " ", match.group(1)).title()

    head = text.split(",")[0].strip()
    return head.title() if head.isupper() else head


@dataclass
class Dig:
    """One anomaly to be staked, plus everything derived for it."""

    # --- straight off the dig sheet -------------------------------------
    name: str = ""
    source_file: str = ""
    odometer: Optional[float] = None
    station_ft: Optional[int] = None
    us_weld_distance: Optional[float] = None
    ds_weld_distance: Optional[float] = None
    us_agm_ref: Optional[str] = None
    us_agm_distance: Optional[float] = None
    ds_agm_ref: Optional[str] = None
    ds_agm_distance: Optional[float] = None
    ili_latitude: Optional[float] = None
    ili_longitude: Optional[float] = None
    hca_from_digsheet: Optional[str] = None
    feature_description: Optional[str] = None

    # True when AS stationing rises with absolute odometer on this line.
    station_ascends: bool = True

    # --- derived from the alignment sheets ------------------------------
    line_name: str = UNKNOWN
    alignment_sheet: str = UNKNOWN
    tract_number: str = UNKNOWN
    legal_description: str = UNKNOWN
    county: str = UNKNOWN
    state: str = UNKNOWN
    hca: str = UNKNOWN

    # --- filled in by the app / by hand ---------------------------------
    directions: str = ""
    staking_notes: str = ""
    aerial_png: Optional[bytes] = None
    notes: str = ""

    warnings: list = field(default_factory=list)

    # --- convenience ----------------------------------------------------
    @property
    def upstream_reference(self) -> str:
        """A5 on the staking report."""
        name = agm_name(self.us_agm_ref)
        if not name:
            return "Launch Valve"
        return name

    @property
    def downstream_reference(self) -> str:
        """N5 on the staking report."""
        name = agm_name(self.ds_agm_ref)
        if not name:
            return "Receive Valve"
        return name

    @property
    def upstream_feet_to_agm(self) -> Optional[float]:
        """A6. With no upstream AGM the distance is simply the odometer."""
        if self.us_agm_ref:
            return self.us_agm_distance
        return self.odometer

    @property
    def downstream_feet_to_agm(self) -> Optional[float]:
        return self.ds_agm_distance

    @property
    def nearest_agm_label(self) -> str:
        """The cheat sheet's 'Feet from AGM U/S-D/S' cell, e.g. "AGM 520 468.12'"."""
        us, ds = self.us_agm_distance, self.ds_agm_distance
        candidates = []
        if us is not None:
            candidates.append((us, self.upstream_reference))
        if ds is not None:
            candidates.append((ds, self.downstream_reference))
        if not candidates:
            return ""
        distance, label = min(candidates, key=lambda pair: pair[0])
        return f"{label} {distance:.2f}'"

    @property
    def output_basename(self) -> str:
        return f"{self.name}_Staking Report"


# ==========================================================================
# XLSMPATCH
# ==========================================================================
# Edit an .xlsm in place at the package level.
#
# openpyxl cannot round-trip this template: it drops DrawingML shapes, which
# would take the macro-linked Save and PDF buttons with them. So the workbook is
# treated as what it is - a zip of XML parts - and only the bytes that need to
# change are changed. Styles, macros, buttons, print settings and the logo all
# survive untouched.

R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

EXCEL_EPOCH = _dt.datetime(1899, 12, 30)


@dataclass
class ImageSlot:
    """A picture already anchored in the template, ready to be swapped out."""

    media_path: str
    extension: str
    width_emu: int
    height_emu: int

    @property
    def aspect(self) -> float:
        if not self.height_emu:
            return 2.0
        return self.width_emu / self.height_emu


class XlsmPatcher:
    def __init__(self, data: bytes):
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            self.names = list(archive.namelist())
            self.parts = {name: archive.read(name) for name in self.names}
            self.infos = {info.filename: info for info in archive.infolist()}
        self._sheet_paths = self._map_sheets()

    # -- structure ------------------------------------------------------
    def _text(self, path: str) -> str:
        return self.parts[path].decode("utf-8")

    def _map_sheets(self) -> dict:
        workbook = self._text("xl/workbook.xml")
        rels = self._text("xl/_rels/workbook.xml.rels")

        targets = {
            match.group(1): match.group(2)
            for match in re.finditer(
                r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels
            )
        }
        mapping = {}
        for match in re.finditer(r"<sheet\b[^>]*/?>", workbook):
            tag = match.group(0)
            name = re.search(r'name="([^"]*)"', tag)
            rid = re.search(r'r:id="([^"]*)"', tag)
            if not name or not rid:
                continue
            target = targets.get(rid.group(1), "")
            if not target:
                continue
            path = target if target.startswith("xl/") else "xl/" + target.lstrip("/")
            path = path.replace("/./", "/")
            mapping[name.group(1)] = path
        return mapping

    def sheet_names(self) -> list:
        return list(self._sheet_paths)

    # -- cell values ----------------------------------------------------
    def set_values(self, sheet_name: str, values: dict) -> list:
        """Write {'C19': value} into a sheet. Returns cells it could not place."""
        path = self._sheet_paths.get(sheet_name)
        if path is None:
            raise KeyError(f"No sheet named {sheet_name!r} in this workbook")

        xml = self._text(path)
        missing = []
        for ref, value in values.items():
            if value is None:
                continue
            xml, ok = _write_cell(xml, ref, value)
            if not ok:
                missing.append(ref)
        self.parts[path] = xml.encode("utf-8")
        return missing

    # -- images ---------------------------------------------------------
    def find_image_slot(self, sheet_name: str) -> Optional[ImageSlot]:
        """The largest picture anchored on a sheet - the aerial image slot."""
        drawing = self._drawing_for(sheet_name)
        if drawing is None:
            return None
        drawing_xml = self._text(drawing)
        rels_path = _rels_path(drawing)
        if rels_path not in self.parts:
            return None
        rels = self._text(rels_path)
        targets = {
            match.group(1): match.group(2)
            for match in re.finditer(
                r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels
            )
        }

        best = None
        for block in re.finditer(r"<xdr:pic\b.*?</xdr:pic>", drawing_xml, re.S):
            chunk = block.group(0)
            embed = re.search(r'r:embed="([^"]+)"', chunk)
            extent = re.search(r'<a:ext\s+cx="(\d+)"\s+cy="(\d+)"', chunk)
            if not embed or not extent:
                continue
            width, height = int(extent.group(1)), int(extent.group(2))
            target = targets.get(embed.group(1))
            if not target:
                continue
            media = _resolve(drawing, target)
            if media not in self.parts:
                continue
            if best is None or width * height > best.width_emu * best.height_emu:
                best = ImageSlot(
                    media_path=media,
                    extension=media.rsplit(".", 1)[-1].lower(),
                    width_emu=width,
                    height_emu=height,
                )
        return best

    def replace_image(self, slot: ImageSlot, image_bytes: bytes) -> None:
        self.parts[slot.media_path] = image_bytes

    def _drawing_for(self, sheet_name: str) -> Optional[str]:
        path = self._sheet_paths.get(sheet_name)
        if path is None:
            return None
        rels_path = _rels_path(path)
        if rels_path not in self.parts:
            return None
        rels = self._text(rels_path)
        match = re.search(
            r'<Relationship[^>]*Type="[^"]*/drawing"[^>]*Target="([^"]+)"', rels
        )
        if not match:
            match = re.search(
                r'<Relationship[^>]*Target="([^"]+)"[^>]*Type="[^"]*/drawing"', rels
            )
        if not match:
            return None
        return _resolve(path, match.group(1))

    # -- output ---------------------------------------------------------
    def to_bytes(self) -> bytes:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for name in self.names:
                info = self.infos[name]
                new = zipfile.ZipInfo(name, date_time=info.date_time)
                new.compress_type = zipfile.ZIP_DEFLATED
                new.external_attr = info.external_attr
                archive.writestr(new, self.parts[name])
        return buffer.getvalue()


# ---------------------------------------------------------------------------
# Cell writing
# ---------------------------------------------------------------------------

def _cell_pattern(ref: str) -> re.Pattern:
    return re.compile(
        rf'<c r="{ref}"(?P<attrs>[^>]*?)(?:/>|>(?P<body>.*?)</c>)', re.S
    )


def _style_of(attrs: str) -> str:
    match = re.search(r'\ss="(\d+)"', attrs or "")
    return f' s="{match.group(1)}"' if match else ""


def _render(ref: str, style: str, value) -> str:
    if isinstance(value, bool):
        return f'<c r="{ref}"{style} t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float)):
        return f'<c r="{ref}"{style}><v>{value!r}</v></c>'
    if isinstance(value, _dt.datetime):
        serial = (value - EXCEL_EPOCH).total_seconds() / 86400.0
        return f'<c r="{ref}"{style}><v>{serial!r}</v></c>'
    if isinstance(value, _dt.date):
        serial = (_dt.datetime(value.year, value.month, value.day) - EXCEL_EPOCH).days
        return f'<c r="{ref}"{style}><v>{serial}</v></c>'

    text = str(value)
    if text.startswith("="):
        return f'<c r="{ref}"{style}><f>{escape(text[1:])}</f></c>'
    return (
        f'<c r="{ref}"{style} t="inlineStr">'
        f'<is><t xml:space="preserve">{escape(text)}</t></is></c>'
    )


def _write_cell(xml: str, ref: str, value) -> tuple:
    match = _cell_pattern(ref).search(xml)
    if match:
        replacement = _render(ref, _style_of(match.group("attrs")), value)
        return xml[: match.start()] + replacement + xml[match.end():], True

    inserted = _insert_cell(xml, ref, value)
    return (inserted, True) if inserted is not None else (xml, False)


_COLUMN_RE = re.compile(r"([A-Z]+)(\d+)")


def _column_index(letters: str) -> int:
    index = 0
    for character in letters:
        index = index * 26 + (ord(character) - 64)
    return index


def _insert_cell(xml: str, ref: str, value) -> Optional[str]:
    """Add a cell to an existing row, keeping columns in order."""
    parsed = _COLUMN_RE.fullmatch(ref)
    if not parsed:
        return None
    letters, number = parsed.group(1), parsed.group(2)
    target = _column_index(letters)

    row = re.search(rf'<row[^>]*\br="{number}"[^>]*>(.*?)</row>', xml, re.S)
    if not row:
        return None

    body = row.group(1)
    style = ""
    insert_at = len(body)
    for cell in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>.*?</c>)', body, re.S):
        index = _column_index(cell.group(1))
        if index < target:
            style = _style_of(cell.group(2)) or style
        else:
            insert_at = cell.start()
            break

    new_body = body[:insert_at] + _render(ref, style, value) + body[insert_at:]
    return xml[: row.start(1)] + new_body + xml[row.end(1):]


# ---------------------------------------------------------------------------
# Package path helpers
# ---------------------------------------------------------------------------

def _rels_path(part: str) -> str:
    folder, _, name = part.rpartition("/")
    return f"{folder}/_rels/{name}.rels"


def _resolve(base_part: str, target: str) -> str:
    folder = base_part.rpartition("/")[0]
    parts = folder.split("/") if folder else []
    for chunk in target.split("/"):
        if chunk in ("", "."):
            continue
        if chunk == "..":
            if parts:
                parts.pop()
        else:
            parts.append(chunk)
    return "/".join(parts)


# ==========================================================================
# DIGSHEET
# ==========================================================================
# Read anomaly rows out of ONEOK dig sheets (.xlsx / .xlsm / .pdf).
#
# A dig sheet is a full ILI feature listing for a stretch of pipe. Exactly one
# row per dig carries a value in the ``Dig Number`` column - that row is the
# anomaly being staked, and every field the staking report needs comes from it.

# Canonical column names, and the loose spellings seen in the wild.
COLUMN_ALIASES = {
    "dig_number": ["dig number", "dig name", "dig no", "dig #"],
    "feature_id": ["ili feature id", "feature id"],
    "feature_desc": ["feature desc", "feature description"],
    "odometer": ["absolute odometer", "abs odometer", "odometer", "odo"],
    "station": ["as stationing", "stationing", "station"],
    "depth": ["depth"],
    "wall_thickness": ["measured wall thickness", "wall thickness"],
    "smys": ["smys"],
    "length": ["length"],
    "width": ["width"],
    "oclock": ["feature o'clock", "feature oclock", "o'clock"],
    "mop": ["mop"],
    "us_weld_distance": ["us weld distance", "u/s weld distance"],
    "ds_weld_distance": ["ds weld distance", "d/s weld distance"],
    "us_agm_ref": ["us agm ref", "u/s agm ref", "us agm reference"],
    "us_agm_distance": ["us agm distance", "u/s agm distance"],
    "ds_agm_ref": ["ds agm ref", "d/s agm ref", "ds agm reference"],
    "ds_agm_distance": ["ds agm distance", "d/s agm distance"],
    "joint_length": ["joint length"],
    "joint_number": ["joint number"],
    "comments": ["data comments", "comments"],
    "latitude": ["ili latitude", "latitude", "lat"],
    "longitude": ["ili longitude", "longitude", "long", "lon"],
    "is_hca": ["is hca", "hca"],
    "flag": ["flag"],
    "vendor_called": ["vendor called previous dig"],
}

_ALIAS_LOOKUP = {
    alias: canonical
    for canonical, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}


def _normalise_header(text) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()


def _canonical(header) -> Optional[str]:
    return _ALIAS_LOOKUP.get(_normalise_header(header))


def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("'", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _clean(value) -> Optional[str]:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


# ---------------------------------------------------------------------------
# Stationing direction
# ---------------------------------------------------------------------------

def detect_station_ascends(rows: Iterable[dict]) -> bool:
    """Does AS stationing rise with absolute odometer on this line?

    This decides the sign of the cheat sheet's weld formulas, so it is derived
    from the sheet's own data rather than guessed. Ties default to ascending.
    """
    points = []
    for row in rows:
        odo = _to_float(row.get("odometer"))
        sta = station_to_feet(row.get("station"))
        if odo is not None and sta is not None:
            points.append((odo, sta))
    if len(points) < 2:
        return True
    points.sort(key=lambda pair: pair[0])
    rises = sum(1 for (_, a), (_, b) in zip(points, points[1:]) if b > a)
    falls = sum(1 for (_, a), (_, b) in zip(points, points[1:]) if b < a)
    return rises >= falls


# ---------------------------------------------------------------------------
# Row -> Dig
# ---------------------------------------------------------------------------

def _row_to_dig(row: dict, source_file: str, ascends: bool) -> Dig:
    hca_raw = _clean(row.get("is_hca"))
    hca = None
    if hca_raw:
        lowered = hca_raw.lower()
        if lowered in {"y", "yes", "true", "1"}:
            hca = "Yes"
        elif lowered in {"n", "no", "false", "0"}:
            hca = "No"

    dig = Dig(
        name=_clean(row.get("dig_number")) or "",
        source_file=source_file,
        odometer=_to_float(row.get("odometer")),
        station_ft=station_to_feet(row.get("station")),
        us_weld_distance=_to_float(row.get("us_weld_distance")),
        ds_weld_distance=_to_float(row.get("ds_weld_distance")),
        us_agm_ref=_clean(row.get("us_agm_ref")),
        us_agm_distance=_to_float(row.get("us_agm_distance")),
        ds_agm_ref=_clean(row.get("ds_agm_ref")),
        ds_agm_distance=_to_float(row.get("ds_agm_distance")),
        ili_latitude=_to_float(row.get("latitude")),
        ili_longitude=_to_float(row.get("longitude")),
        hca_from_digsheet=hca,
        feature_description=_clean(row.get("comments")),
        station_ascends=ascends,
    )
    if dig.station_ft is None:
        dig.warnings.append("No AS stationing found on the anomaly row.")
    if dig.ili_latitude is None or dig.ili_longitude is None:
        dig.warnings.append("No ILI latitude/longitude - aerial image cannot be generated.")
    return dig


# ---------------------------------------------------------------------------
# Excel dig sheets
# ---------------------------------------------------------------------------

def parse_excel_digsheet(data: bytes, source_file: str) -> list[Dig]:
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    digs: list[Dig] = []

    for sheet in workbook.worksheets:
        grid = list(sheet.iter_rows(values_only=True))
        header_index, mapping = _find_excel_header(grid)
        if mapping is None:
            continue

        rows = []
        for raw in grid[header_index + 1:]:
            row = {}
            for index, canonical in mapping.items():
                if index < len(raw):
                    row[canonical] = raw[index]
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)

        ascends = detect_station_ascends(rows)
        for row in rows:
            if _clean(row.get("dig_number")):
                digs.append(_row_to_dig(row, source_file, ascends))

    workbook.close()
    return digs


def _find_excel_header(grid) -> tuple[int, Optional[dict]]:
    """Locate the header row and map column index -> canonical name."""
    for index, raw in enumerate(grid[:25]):
        mapping = {}
        for position, cell in enumerate(raw or ()):
            canonical = _canonical(cell)
            if canonical:
                mapping[position] = canonical
        if "odometer" in mapping.values() and "dig_number" in mapping.values():
            return index, mapping
    return -1, None


# ---------------------------------------------------------------------------
# PDF dig sheets
# ---------------------------------------------------------------------------

_HEADER_TOKENS = {
    "dig", "number", "ili", "feature", "id", "desc", "absolute", "odometer",
    "as", "stationing", "depth", "measured", "wall", "thickness", "smys",
    "length", "width", "o'clock", "mop", "us", "ds", "weld", "distance",
    "agm", "ref", "joint", "data", "comments", "latitude", "longitude",
    "flag", "is", "hca", "vendor", "called", "previous",
}

# Dig names look like NL3DH-24-F1, H64M-26-F11, STSB-24-F1, RJSJ-25-F1.
DIG_NAME_RE = re.compile(r"\b([A-Z0-9]{2,10}-\d{2}-F\d{1,3})\b")

# Anchored at the start of a cell, and guarded against running into the
# feature id that follows it ("NL3DH-24-F9" then "40116291").
DIG_NAME_LEAD_RE = re.compile(r"^([A-Z0-9]{2,10}-\d{2}-F\d{1,3})(?![0-9])")

STATION_TOKEN_RE = re.compile(r"\b\d{1,5}\+\d{2}\b")
TWO_DP_RE = re.compile(r"\b\d{1,7}\.\d{2}\b")
ANY_FLOAT_RE = re.compile(r"\b\d{1,7}\.\d{1,3}\b")
LATLON_TAIL_RE = re.compile(r"(-?\d{1,3}\.\d{4,})\s+(-?\d{1,3}\.\d{4,})\s*$")
AGM_TOKEN_RE = re.compile(r"\bAGM\s+\d+\b")
# Both reference cells carry a station, whether they name an AGM, a launch or
# receive valve, or anything else. "Sta." is the reliable anchor; "AGM" is not.
STA_ANCHOR_RE = re.compile(r"\bSta\.")


def parse_pdf_digsheet(data: bytes, source_file: str) -> list[Dig]:
    """Find the anomaly row in a PDF dig sheet.

    Three independent signals identify it, in order of confidence:

      1. a value in the Dig Number column,
      2. the dig name from the heading at the top centre of page one
         appearing in the row,
      3. the row sitting inside a yellow highlight band.

    Field values are read out of the row text with anchored patterns rather
    than by trusting column detection, because the columns are the fragile
    part: the header wraps over several lines and merges unpredictably.
    Detected columns are still used to fill anything the patterns miss.
    """
    rows, heading, bands, columns = _read_pdf(data)

    # With columns, the cells are authoritative. The text patterns are only a
    # fallback for a sheet whose header could not be read - running them over
    # a positionally-sorted row would reintroduce the interleaving problem.
    if not columns:
        for row in rows:
            parsed = _parse_row_text(row["_raw"])
            for key, value in parsed.items():
                if value:
                    row[key] = value

    anomalies, _ = _select_anomaly_rows(rows, heading, bands)
    if not anomalies:
        return []

    ascends = detect_station_ascends(rows)
    digs = []
    for row, name in anomalies:
        row = dict(row)
        row["dig_number"] = name
        digs.append(_row_to_dig(row, source_file, ascends))
    return digs


def _read_pdf(data: bytes):
    """Pull rows, the page-one heading, highlight bands and columns out of a PDF.

    Cells are recovered from characters in **content-stream order**, not by
    sorting on position. These sheets are printed from Excel with the text
    overflowing its columns, so a reference cell and the number beside it
    physically overlap in x - sorting by position interleaves them into
    nonsense like "Dan4v9il4le7,." Each cell, however, is emitted as one
    contiguous run, so a run ends where x jumps backwards or leaves a gap.
    """
    import pdfplumber

    rows: list[dict] = []
    bands: list[tuple] = []
    heading = None
    columns = None

    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for index, page in enumerate(pdf.pages):
            lines = _char_lines(page)
            if not lines:
                continue
            if heading is None:
                heading = _page_heading_dig_name(
                    page, page.extract_words(keep_blank_chars=False)
                )
            found = _pdf_columns(lines)
            if found:
                columns = found
            for top, bottom in _highlight_bands(page):
                bands.append((index, top, bottom))
            rows.extend(_pdf_rows(lines, columns, index))

    return rows, heading, bands, columns


def _char_lines(page, tolerance: float = 2.0):
    """Group characters into visual lines, keeping content-stream order."""
    buckets: dict = {}
    for char in page.chars:
        key = None
        for existing in buckets:
            if abs(existing - char["top"]) <= tolerance:
                key = existing
                break
        if key is None:
            key = char["top"]
            buckets[key] = []
        buckets[key].append(char)
    return [(top, buckets[top]) for top in sorted(buckets)]


def _segment_cells(chars, backwards: float = 1.0, gap: float = 2.0):
    """Split one line's characters into cells at run boundaries."""
    cells = []
    current = []
    previous = None
    for char in chars:
        if previous is not None and (
            char["x0"] < previous["x1"] - backwards
            or char["x0"] > previous["x1"] + gap
        ):
            cells.append(_finish_cell(current))
            current = []
        current.append(char)
        previous = char
    if current:
        cells.append(_finish_cell(current))
    return [cell for cell in cells if cell[2]]


def _finish_cell(chars):
    return (
        min(c["x0"] for c in chars),
        max(c["x1"] for c in chars),
        "".join(c["text"] for c in chars).strip(),
    )


def _pdf_columns(lines):
    """Column start positions and names, read off the wrapped header block.

    Header text wraps over two printed lines ("US AGM" above "Distance").
    Both sit at the same x, so the pieces are grouped by their start position
    and joined top to bottom.
    """
    header_lines = []
    for top, chars in lines[:20]:
        cells = _segment_cells(chars)
        if not cells:
            continue
        words = " ".join(text for _, _, text in cells).lower().split()
        hits = sum(1 for word in words if word.strip(":'") in _HEADER_TOKENS)
        if words and hits >= max(3, len(words) * 0.6):
            header_lines.append((top, cells))
    if not header_lines:
        return None

    groups: dict = {}
    for top, cells in header_lines:
        for x0, _, text in cells:
            key = None
            for existing in groups:
                if abs(existing - x0) <= 1.5:
                    key = existing
                    break
            if key is None:
                key = x0
                groups[key] = []
            groups[key].append((top, text))

    columns = []
    for x0 in sorted(groups):
        parts = [text for _, text in sorted(groups[x0])]
        name = _canonical(" ".join(parts))
        if name:
            columns.append((x0, name))
    return columns if len(columns) >= 6 else None


def _pdf_rows(lines, columns, page_index: int):
    """Turn each line's cells into a row keyed by canonical column name."""
    rows = []
    for top, chars in lines:
        cells = _segment_cells(chars)
        if not cells:
            continue

        words = " ".join(text for _, _, text in cells).lower().split()
        hits = sum(1 for word in words if word.strip(":'") in _HEADER_TOKENS)
        if words and hits >= max(3, len(words) * 0.6):
            continue  # a header line

        row: dict = {}
        if columns:
            for x0, _, text in cells:
                name = None
                for start, candidate in columns:
                    if x0 >= start - 1.0:
                        name = candidate
                    else:
                        break
                if name is None:
                    continue
                row[name] = f"{row[name]} {text}".strip() if name in row else text

        ordered = sorted(cells, key=lambda cell: cell[0])
        row["_raw"] = " ".join(text for _, _, text in ordered)
        row["_top"] = top
        row["_page"] = page_index
        row["_lead_words"] = [text for _, _, text in ordered[:2]]

        if (
            row.get("station")
            or row.get("odometer")
            or STATION_TOKEN_RE.search(row["_raw"])
        ):
            rows.append(row)
    return rows


def _page_heading_dig_name(page, words) -> Optional[str]:
    """The dig name printed at the top centre of the first page."""
    zone = page.height * 0.20
    centre = page.width / 2.0
    best, best_distance = None, None

    for word in words:
        if word["top"] > zone:
            continue
        match = DIG_NAME_RE.search(word["text"].replace(" ", "").upper())
        if not match:
            continue
        distance = abs((word["x0"] + word["x1"]) / 2.0 - centre)
        if best_distance is None or distance < best_distance:
            best, best_distance = match.group(1), distance
    if best:
        return best

    # The name may be split across words; try the top zone joined up.
    joined = "".join(w["text"] for w in words if w["top"] <= zone).upper()
    match = DIG_NAME_RE.search(joined)
    return match.group(1) if match else None


def _is_yellowish(colour) -> bool:
    if not colour:
        return False
    try:
        values = tuple(float(component) for component in colour)
    except (TypeError, ValueError):
        return False
    if len(values) == 3:
        red, green, blue = values
        return red > 0.70 and green > 0.60 and blue < 0.60 and abs(red - green) < 0.40
    if len(values) == 4:
        cyan, magenta, yellow, black = values
        return yellow > 0.35 and cyan < 0.35 and magenta < 0.45 and black < 0.35
    if len(values) == 1:
        return False
    return False


def _highlight_bands(page) -> list:
    """Vertical bands covered by yellow fills - the highlighted anomaly row."""
    raw = []
    for shape in list(page.rects) + list(page.curves):
        if shape.get("fill") is False:
            continue
        if not _is_yellowish(shape.get("non_stroking_color")):
            continue
        height = shape["bottom"] - shape["top"]
        width = shape["x1"] - shape["x0"]
        if height <= 0 or height > page.height * 0.12 or width < 3:
            continue
        raw.append((shape["top"], shape["bottom"]))

    if not raw:
        return []
    raw.sort()
    merged = [list(raw[0])]
    for top, bottom in raw[1:]:
        if top <= merged[-1][1] + 1.5:
            merged[-1][1] = max(merged[-1][1], bottom)
        else:
            merged.append([top, bottom])
    return [tuple(band) for band in merged]


def _select_anomaly_rows(rows, heading, bands):
    """The anomaly row must satisfy all three conditions at once.

    A dig sheet can carry other highlighted rows that are not the call
    anomaly, so no single signal is trusted on its own. A row qualifies only
    when it:

      1. begins with a dig name - that is, has a value in the Dig Number
         column, which is the leftmost column;
      2. carries the same dig name as the heading at the top centre of page
         one; and
      3. sits inside a yellow highlight band.

    Returns ``(matches, tally)``. The tally records how many rows survived
    each condition so a failure can say which one eliminated everything
    rather than just reporting nothing found.
    """
    needle = heading.replace(" ", "").upper() if heading else None
    tally = {
        "rows": len(rows),
        "start_with_a_dig_name": 0,
        "and_match_the_heading": 0,
        "and_are_highlighted": 0,
        "heading": heading,
        "dig_names_seen": [],
    }

    matches = []
    for row in rows:
        lead = _leading_dig_name(row)
        if not lead:
            continue
        tally["start_with_a_dig_name"] += 1
        if lead not in tally["dig_names_seen"]:
            tally["dig_names_seen"].append(lead)

        if not needle or lead.replace(" ", "").upper() != needle:
            continue
        tally["and_match_the_heading"] += 1

        if not _in_band(row, bands):
            continue
        tally["and_are_highlighted"] += 1

        matches.append((row, lead))

    return matches, tally


def _anomaly_failure_reason(tally, bands) -> str:
    """Say which of the three conditions eliminated every row."""
    if not tally["rows"]:
        return "no data rows were found in the PDF at all."
    if not tally["start_with_a_dig_name"]:
        return (
            "no row starts with a dig name, so no row has a value in the "
            "Dig Number column."
        )
    if not tally["heading"]:
        return (
            "no dig name was found in the heading at the top centre of page "
            f"one, though {tally['start_with_a_dig_name']} row(s) carry a Dig "
            f"Number ({', '.join(tally['dig_names_seen'])})."
        )
    if not tally["and_match_the_heading"]:
        return (
            f"the heading reads {tally['heading']}, but the Dig Number "
            f"column holds {', '.join(tally['dig_names_seen'])} - they do "
            "not match."
        )
    if not bands:
        return (
            f"the row for {tally['heading']} was found, but no yellow "
            "highlight was detected anywhere in the PDF."
        )
    return (
        f"the row for {tally['heading']} was found and the PDF has "
        f"{len(bands)} highlighted band(s), but that row is not inside one."
    )


def _leading_dig_name(row) -> Optional[str]:
    """The dig name when the row carries a Dig Number."""
    cell = _clean(row.get("dig_number"))
    if cell:
        match = DIG_NAME_LEAD_RE.match(cell.replace(" ", "").upper())
        if match:
            return match.group(1)
        return None

    words = row.get("_lead_words") or []
    if not words:
        return None
    first = words[0].replace(" ", "").upper()
    match = DIG_NAME_LEAD_RE.match(first)
    if match:
        return match.group(1)
    if len(words) > 1:
        # The name can be split across two words ("NL3DH-24-" then "F9").
        joined = (words[0] + words[1]).replace(" ", "").upper()
        match = DIG_NAME_LEAD_RE.match(joined)
        if match:
            return match.group(1)
    return None


def _in_band(row, bands) -> bool:
    for page_index, top, bottom in bands:
        if page_index != row.get("_page"):
            continue
        if top - 2.0 <= row.get("_top", -1) <= bottom + 2.0:
            return True
    return False


def _parse_row_text(text: str) -> dict:
    """Read a row's values straight off its text, anchored on stable landmarks.

    The column order is fixed, so each field can be found relative to something
    unmistakable: the station is the only ``nnnn+nn`` token, the odometer is the
    float before it, the weld distances are the last two two-decimal numbers
    before the first AGM reference, and so on.
    """
    out: dict = {}
    line = " ".join(text.split())
    if not line:
        return out

    match = DIG_NAME_RE.search(line.replace(" ", "").upper())
    if match:
        out["dig_number_guess"] = match.group(1)

    tail = LATLON_TAIL_RE.search(line)
    if tail:
        out["latitude"], out["longitude"] = tail.group(1), tail.group(2)

    station = STATION_TOKEN_RE.search(line)
    search_from = 0
    if station:
        out["station"] = station.group(0)
        search_from = station.end()
        before = line[: station.start()]
        floats = ANY_FLOAT_RE.findall(before)
        if floats:
            out["odometer"] = floats[-1]

    _read_references(line, out, search_from)

    return out


TOKEN_2DP_RE = re.compile(r"^\d{1,7}\.\d{2}$")
HAS_LETTER_RE = re.compile(r"[A-Za-z]")


def _read_references(line: str, out: dict, search_from: int = 0) -> None:
    """Read the two reference cells and the distances that follow them.

    Column order is: ... MOP, US Weld Distance, DS Weld Distance,
    US AGM Ref, US AGM Distance, DS AGM Ref, DS AGM Distance, Joint Length ...

    Everything between the stationing and the first reference is numeric -
    depth, wall thickness, SMYS, length, width, o'clock, MOP and the two weld
    distances - so **the first token containing a letter after the station
    begins the upstream reference**. That is the anchor.

    Anchoring on "AGM" is wrong: an upstream reference is often "Launch Valve".
    Anchoring on "Sta." is also wrong: a launch or receive valve reference does
    not always carry a station, and when it does not, nothing parses and the
    downstream reference silently falls back to a default.

    Each reference then runs up to the next standalone two-decimal number,
    which is its distance.
    """
    tokens = [(m.start(), m.end(), m.group(0)) for m in re.finditer(r"\S+", line)]
    if not tokens:
        return

    def first_word_index(after_position: int) -> Optional[int]:
        for index, (begin, _, text) in enumerate(tokens):
            if begin >= after_position and HAS_LETTER_RE.search(text):
                return index
        return None

    def first_distance_index(from_index: int) -> Optional[int]:
        for index in range(from_index, len(tokens)):
            if TOKEN_2DP_RE.match(tokens[index][2]):
                return index
        return None

    us_start = first_word_index(search_from)
    if us_start is None:
        return
    us_distance = first_distance_index(us_start)
    if us_distance is None:
        return

    # The two standalone two-decimal numbers before the reference are the
    # weld distances.
    preceding = [
        text for begin, _, text in tokens[:us_start]
        if TOKEN_2DP_RE.match(text)
    ]
    if len(preceding) >= 2:
        out["us_weld_distance"] = preceding[-2]
        out["ds_weld_distance"] = preceding[-1]

    reference = line[tokens[us_start][0]:tokens[us_distance][0]].strip(" ,")
    if reference:
        out["us_agm_ref"] = reference
    out["us_agm_distance"] = tokens[us_distance][2]

    ds_start = first_word_index(tokens[us_distance][1])
    if ds_start is None:
        return
    ds_distance = first_distance_index(ds_start)
    if ds_distance is None:
        return

    reference = line[tokens[ds_start][0]:tokens[ds_distance][0]].strip(" ,")
    if reference:
        out["ds_agm_ref"] = reference
    out["ds_agm_distance"] = tokens[ds_distance][2]


def _cluster_lines(words, tolerance: float = 2.5):
    """Group words into visual lines by their vertical position."""
    lines: list[list[dict]] = []
    for word in sorted(words, key=lambda w: (round(w["top"], 1), w["x0"])):
        if lines and abs(word["top"] - lines[-1][0]["top"]) <= tolerance:
            lines[-1].append(word)
        else:
            lines.append([word])
    for line in lines:
        line.sort(key=lambda w: w["x0"])
    return lines


def _looks_like_header(tokens) -> bool:
    if not tokens:
        return False
    hits = sum(1 for token in tokens if token in _HEADER_TOKENS)
    return hits >= max(3, len(tokens) * 0.6)


def pdf_digsheet_report(data: bytes, filename: str) -> dict:
    """What the parser saw in a PDF, for when it finds nothing."""
    try:
        rows, heading, bands, columns = _read_pdf(data)
    except Exception as error:  # noqa: BLE001
        return {"file": filename, "error": str(error)}

    matches, tally = _select_anomaly_rows(rows, heading, bands)

    samples = []
    for row in rows[:400]:
        if _leading_dig_name(row):
            samples.append(row["_raw"][:180])
        if len(samples) >= 5:
            break

    return {
        "file": filename,
        "why_nothing_matched": _anomaly_failure_reason(tally, bands),
        "heading_at_top_of_page_one": heading,
        "rows_read": tally["rows"],
        "1_start_with_a_dig_name": tally["start_with_a_dig_name"],
        "2_and_match_the_heading": tally["and_match_the_heading"],
        "3_and_are_highlighted": tally["and_are_highlighted"],
        "dig_names_in_the_dig_number_column": tally["dig_names_seen"],
        "highlight_bands_found": len(bands),
        "columns_detected": [name for _, _, name in columns] if columns else None,
        "example_rows_with_a_dig_number": samples,
        "first_data_row": rows[0]["_raw"][:180] if rows else None,
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_digsheet(data: bytes, filename: str) -> list[Dig]:
    lowered = filename.lower()
    if lowered.endswith(".pdf"):
        return parse_pdf_digsheet(data, filename)
    if lowered.endswith((".xlsx", ".xlsm", ".xls")):
        return parse_excel_digsheet(data, filename)
    raise ValueError(f"Unsupported dig sheet format: {filename}")


# ==========================================================================
# ALIGNMENT
# ==========================================================================
# Pull route, PLSS, county and HCA information out of alignment sheet PDFs.
#
# ONEOK alignment sheets carry a real text layer, so the title block fields come
# out reliably. The banded information (PLSS sections, HCA) is positional: the
# plan view runs linearly left-to-right along stationing, so a station tick
# regression turns an x coordinate into a station and back again.
#
# Anything that cannot be established with confidence is left as ``Unknown``
# rather than guessed - a wrong tract number in a signed report is worse than a
# blank one.

STATION_RE = re.compile(r"\b(\d{1,5})\+(\d{2})\b")
RANGE_RE = re.compile(r"From\s+(\d{1,5}\+\d{2})\s+To\s+(\d{1,5}\+\d{2})", re.I)
PIPELINE_RE = re.compile(r"Pipeline\s+Number:?\s*(\d+)", re.I)
SHEET_RE = re.compile(r"Sheet\s+(\d+)\s+of\s+(\d+)", re.I)
COUNTY_RE = re.compile(r"([A-Z][A-Z .'\-]{1,30}?)\s+COUNTY\b\s*,\s*([A-Z][A-Z ]+)")
# The sheets carry a "COUNTY" row label as well as the county name, and the two
# can extract onto one line.
FILENAME_SHEET_RE = re.compile(r"(\d{4,6})[_\-](\d{1,3})\b")
PLSS_RE = re.compile(r"T\s*(\d+)\s*([NS])\s*,\s*R\s*(\d+)\s*([EW])\s*,\s*SEC\s*(\d+)", re.I)
ROUTE_RE = re.compile(r"\((\d{4,6}(?:-\d+)?)\)\s*([A-Za-z].*)")


@dataclass
class BandSegment:
    """One stretch of a banded row, bounded by the dividers drawn around it."""

    label: str
    station_low: Optional[int] = None
    station_high: Optional[int] = None

    def contains(self, station_ft) -> bool:
        if station_ft is None or self.station_low is None:
            return False
        return self.station_low <= station_ft <= self.station_high

    @property
    def midpoint(self) -> Optional[int]:
        if self.station_low is None:
            return None
        return (self.station_low + self.station_high) // 2


@dataclass
class CountySegment(BandSegment):
    county: str = UNKNOWN
    state: str = UNKNOWN


@dataclass
class PlssSegment(BandSegment):
    section: str = ""


@dataclass
class AlignmentSheet:
    filename: str
    pipeline_number: str = UNKNOWN
    sheet_number: str = UNKNOWN
    route_name: str = UNKNOWN
    county: str = UNKNOWN
    state: str = UNKNOWN
    station_start: Optional[int] = None
    station_end: Optional[int] = None
    counties: list = field(default_factory=list)      # list[CountySegment]
    plss: list = field(default_factory=list)          # list[PlssSegment]
    hca_ranges: list = field(default_factory=list)     # list[tuple[int, int]]
    hca_reliable: bool = False
    range_is_inferred: bool = False

    @property
    def sheet_id(self) -> str:
        """The value written into 'Alignment Sheet:' - e.g. 10222_41.

        The filename is preferred: it already carries the form used on the
        reports, whereas the sheet prints a zero-padded number ("Sheet 041").
        """
        stem = self.filename.rsplit("/", 1)[-1]
        stem = re.sub(r"\.pdf$", "", stem, flags=re.I)
        match = FILENAME_SHEET_RE.search(stem)
        if match:
            return f"{match.group(1)}_{match.group(2)}"
        if self.pipeline_number != UNKNOWN and self.sheet_number != UNKNOWN:
            return f"{self.pipeline_number}_{self.sheet_number}"
        stem = re.sub(r"\s*alignment\s*sheet\s*", "", stem, flags=re.I).strip()
        return stem or UNKNOWN

    def describe(self) -> str:
        if self.station_start is None:
            return f"{self.sheet_id} (no station range found)"
        low, high = sorted((self.station_start, self.station_end))
        suffix = " approx" if self.range_is_inferred else ""
        return f"{self.sheet_id} {feet_to_station(low)}-{feet_to_station(high)}{suffix}"

    def covers(self, station_ft: Optional[int]) -> bool:
        if station_ft is None or self.station_start is None or self.station_end is None:
            return False
        low, high = sorted((self.station_start, self.station_end))
        return low <= station_ft <= high

    def county_for(self, station_ft: Optional[int]) -> Optional[CountySegment]:
        return _segment_for(self.counties, station_ft)

    def plss_for(self, station_ft: Optional[int]) -> Optional[PlssSegment]:
        return _segment_for(self.plss, station_ft)

    def hca_for(self, station_ft: Optional[int]) -> str:
        if not self.hca_reliable or station_ft is None:
            return UNKNOWN
        for low, high in self.hca_ranges:
            if low <= station_ft <= high:
                return "Yes"
        return "No"


def _segment_for(segments, station_ft):
    """The band segment covering a station, else the nearest one."""
    if not segments:
        return None
    for segment in segments:
        if segment.contains(station_ft):
            return segment
    if len(segments) == 1:
        return segments[0]
    if station_ft is None:
        return None
    located = [s for s in segments if s.midpoint is not None]
    if not located:
        return None
    return min(located, key=lambda s: abs(s.midpoint - station_ft))


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_alignment_sheet(data: bytes, filename: str) -> AlignmentSheet:
    import pdfplumber

    sheet = AlignmentSheet(filename=filename)

    with pdfplumber.open(io.BytesIO(data)) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ""
        words = page.extract_words(keep_blank_chars=False)

        _read_title_block(sheet, text, words, page.width, page.height)

        axis = _station_axis(words)
        if axis is None:
            return sheet
        anchors, span, plan = axis
        if sheet.station_start is None:
            sheet.station_start, sheet.station_end = span
            sheet.range_is_inferred = True

        bands = _bands(page, words)
        _read_counties(sheet, page, bands, anchors, plan)
        _read_plss(sheet, page, bands, anchors, plan)
        _read_hca(sheet, page, bands, words, anchors, plan)

    return sheet


def _station_axis(words):
    """Fit station = slope * x + intercept from the axis tick labels.

    Returns the fit, the span of stations the sheet covers, and the x extent
    of the plan area. The ticks are the only station labels sharing a
    horizontal line, so the largest such group is the axis - taking the span
    from every station-shaped token on the page instead picks up strays and
    produces a nonsense range.
    """
    ticks = []
    for word in words:
        text = word["text"].strip()
        if STATION_RE.fullmatch(text):
            ticks.append((
                (word["x0"] + word["x1"]) / 2.0, word["top"], station_to_feet(text)
            ))
    if len(ticks) < 3:
        return None

    rows: dict = {}
    for x, top, station in ticks:
        rows.setdefault(int(round(top / 3.0)), []).append((x, station))
    best = max(rows.values(), key=len)
    if len(best) < 3:
        return None

    best.sort()
    xs = [point[0] for point in best]
    ys = [point[1] for point in best]
    count = len(xs)
    mean_x = sum(xs) / count
    mean_y = sum(ys) / count
    denominator = sum((x - mean_x) ** 2 for x in xs)
    if denominator == 0:
        return None
    slope = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys)) / denominator
    intercept = mean_y - slope * mean_x

    residuals = [abs(slope * x + intercept - y) for x, y in zip(xs, ys)]
    spread = max(ys) - min(ys)
    if spread == 0 or max(residuals) > spread * 0.05:
        return None

    return (slope, intercept), (min(ys), max(ys)), (min(xs), max(xs))


def _station_at(anchors, x) -> Optional[int]:
    if not anchors or x is None:
        return None
    slope, intercept = anchors
    return int(round(slope * x + intercept))


def _bands(page, words):
    """The sheet's banded rows, split by the full-width horizontal rules.

    Each band carries its row label down the left edge - COUNTY, PLSS, CLASS,
    HCA - which is what identifies it.
    """
    rules = sorted({
        round(edge["top"], 1)
        for edge in list(page.lines) + list(page.edges)
        if abs(edge["bottom"] - edge["top"]) < 1.5
        and (edge["x1"] - edge["x0"]) > page.width * 0.7
    })

    bands = []
    for top, bottom in zip(rules, rules[1:]):
        if bottom - top < 3:
            continue
        labels = {
            word["text"].strip().upper()
            for word in words
            if word["x0"] < 70 and top - 1 <= word["top"] <= bottom + 1
        }
        bands.append({"top": top, "bottom": bottom, "labels": labels})
    return bands


def _band_named(bands, name: str):
    for band in bands:
        if name in band["labels"]:
            return band
    return None


def _dividers(page, band, plan):
    """The vertical rules that split a band into its segments."""
    height = band["bottom"] - band["top"]
    xs = set()
    for edge in list(page.lines) + list(page.edges):
        if abs(edge["x1"] - edge["x0"]) >= 2:
            continue
        if (edge["bottom"] - edge["top"]) < height * 0.35:
            continue
        if edge["top"] < band["top"] - 1 or edge["bottom"] > band["bottom"] + 1:
            continue
        xs.add(round(edge["x0"], 1))
    return sorted(xs)


def _segment_bounds(dividers, plan):
    """Turn divider positions into (x_low, x_high) spans covering the plan."""
    low, high = plan
    edges = sorted({round(x, 1) for x in dividers})
    edges = [x for x in edges if low - 40 <= x <= high + 40]
    if not edges:
        return [(low, high)]
    if edges[0] > low:
        edges.insert(0, low)
    if edges[-1] < high:
        edges.append(high)
    return list(zip(edges, edges[1:]))


def _station_range(anchors, x_low, x_high):
    a = _station_at(anchors, x_low)
    b = _station_at(anchors, x_high)
    if a is None or b is None:
        return None, None
    return min(a, b), max(a, b)


def _read_title_block(sheet, text, words, page_width, page_height):
    """Route name, pipeline and sheet numbers off the title block.

    The station range is NOT read here. Real sheets do not always carry a
    "From X To Y" line, and the axis ticks give a better answer anyway.
    """
    match = RANGE_RE.search(text)
    if match:
        sheet.station_start = station_to_feet(match.group(1))
        sheet.station_end = station_to_feet(match.group(2))

    match = PIPELINE_RE.search(text)
    if match:
        sheet.pipeline_number = match.group(1)

    match = SHEET_RE.search(text)
    if match:
        sheet.sheet_number = match.group(1)

    sheet.route_name = _find_route_name(text, words, page_width, page_height)


def _find_route_name(text, words, page_width, page_height) -> str:
    """The route name sits in the title block, bottom right of the sheet.

    Found anywhere on the line rather than only at its start - the label
    "Route Name:" and other title block text can share an extracted line with
    the value. Where several candidates exist, the one furthest to the bottom
    right wins, because that is the title block.
    """
    candidates = []
    for raw in text.splitlines():
        match = ROUTE_RE.search(raw)
        if not match:
            continue
        candidate = raw[match.start():].strip()
        # Trim any trailing label that ran onto the same extracted line.
        candidate = re.split(r"\s{3,}", candidate)[0].strip()
        if len(candidate) > 12:
            candidates.append(candidate)

    if not candidates:
        return UNKNOWN
    if len(candidates) == 1:
        return candidates[0]

    best, best_score = candidates[0], -1.0
    for candidate in candidates:
        head = candidate.split()[0]
        for word in words:
            if word["text"].startswith(head[:8]):
                score = (word["x0"] / page_width) + (word["top"] / page_height)
                if score > best_score:
                    best, best_score = candidate, score
    return best


def _state_abbrev(name: str) -> str:
    states = {
        "KANSAS": "KS", "OKLAHOMA": "OK", "TEXAS": "TX", "NEBRASKA": "NE",
        "MISSOURI": "MO", "COLORADO": "CO", "NEW MEXICO": "NM", "ARKANSAS": "AR",
        "IOWA": "IA", "ILLINOIS": "IL", "WYOMING": "WY", "NORTH DAKOTA": "ND",
        "SOUTH DAKOTA": "SD", "MINNESOTA": "MN", "MONTANA": "MT",
        "LOUISIANA": "LA", "MISSISSIPPI": "MS",
    }
    return states.get(name.upper().strip(), name.title())


NAME_WORD_RE = re.compile(r"[A-Z][A-Z.'\-]*$")


def _band_labels(page, band, plan, predicate):
    """Text items sitting inside a band's plan area, with their centres."""
    found = []
    for line in _cluster_lines(
        [w for w in page.extract_words(keep_blank_chars=False)
         if band["top"] - 1 <= w["top"] <= band["bottom"] + 1
         and plan[0] - 40 <= w["x0"] <= plan[1] + 40],
        tolerance=3.0,
    ):
        found.extend(predicate(line))
    return found


def _read_counties(sheet, page, bands, anchors, plan) -> None:
    """The county band - one entry per stretch between the band's dividers."""
    band = _band_named(bands, "COUNTY")
    if band is None:
        return

    def counties_in(line):
        results = []
        for index, word in enumerate(line):
            if word["text"].strip().upper().rstrip(",") != "COUNTY":
                continue
            name_parts, back = [], index - 1
            while back >= 0 and len(name_parts) < 2:
                if line[back + 1]["x0"] - line[back]["x1"] > 14.0:
                    break
                text = line[back]["text"].strip()
                if not NAME_WORD_RE.match(text) or text.upper() == "COUNTY":
                    break
                name_parts.insert(0, text)
                back -= 1
            if not name_parts:
                continue
            state_parts, forward = [], index + 1
            while forward < len(line) and len(state_parts) < 2:
                if line[forward]["x0"] - line[forward - 1]["x1"] > 14.0:
                    break
                text = line[forward]["text"].strip().rstrip(",")
                if not NAME_WORD_RE.match(text) or text.upper() == "COUNTY":
                    break
                state_parts.append(text)
                forward += 1
            if not state_parts:
                continue
            centre = (line[index - len(name_parts)]["x0"]
                      + line[index + len(state_parts)]["x1"]) / 2.0
            results.append((centre, " ".join(name_parts), " ".join(state_parts)))
        return results

    entries = _band_labels(page, band, plan, counties_in)
    if not entries:
        return

    spans = _segment_bounds(_dividers(page, band, plan), plan)
    for centre, name, state in entries:
        low, high = _matching_span(spans, centre, plan)
        station_low, station_high = _station_range(anchors, low, high)
        sheet.counties.append(CountySegment(
            label=f"{name} County",
            station_low=station_low,
            station_high=station_high,
            county=name.title(),
            state=_state_abbrev(state),
        ))

    if sheet.counties:
        sheet.county = sheet.counties[0].county
        sheet.state = sheet.counties[0].state


def _read_plss(sheet, page, bands, anchors, plan) -> None:
    """The PLSS band - each section runs between the dividers drawn around it."""
    band = _band_named(bands, "PLSS")
    if band is None:
        return

    def sections_in(line):
        results = []
        text_parts = [(w["x0"], w["x1"], w["text"]) for w in line]
        joined = " ".join(part[2] for part in text_parts)
        for match in PLSS_RE.finditer(joined):
            township, ns, rng, ew, section = match.groups()
            label = f"T {township}{ns.upper()}, R {rng}{ew.upper()}, SEC {section}"
            # Locate the label by its "SEC <n>" pair.
            for index, word in enumerate(line):
                if word["text"].upper().strip(",") != "SEC":
                    continue
                if index + 1 < len(line) and line[index + 1]["text"].strip().strip(",") == section:
                    centre = (word["x0"] + line[index + 1]["x1"]) / 2.0
                    results.append((centre, label, section))
                    break
        return results

    entries = _band_labels(page, band, plan, sections_in)
    if not entries:
        return

    spans = _segment_bounds(_dividers(page, band, plan), plan)
    seen = set()
    for centre, label, section in entries:
        if label in seen:
            continue
        seen.add(label)
        low, high = _matching_span(spans, centre, plan)
        station_low, station_high = _station_range(anchors, low, high)
        sheet.plss.append(PlssSegment(
            label=label,
            station_low=station_low,
            station_high=station_high,
            section=section,
        ))


def _matching_span(spans, centre, plan):
    for low, high in spans:
        if low <= centre <= high:
            return low, high
    return plan


def _read_hca(sheet, page, bands, words, anchors, plan) -> None:
    """The HCA band.

    HCA stretches are drawn as thin horizontal rules on the HCA row, not as
    filled blocks. When the row is found and carries no rules, the sheet is
    saying there is no HCA here - which is a reliable "No", not "Unknown".
    """
    label = None
    for word in words:
        if word["x0"] < 70 and word["text"].strip().upper() == "HCA":
            label = word
            break
    if label is None:
        return

    # Strictly within the label's own line: the band's border rules sit just
    # outside it and span the full width, and counting one of those as an HCA
    # mark makes the whole sheet look like HCA.
    top = label["top"]
    bottom = label["bottom"]
    plan_width = plan[1] - plan[0]

    ranges = []
    for edge in list(page.lines) + list(page.edges) + list(page.rects):
        if abs(edge["bottom"] - edge["top"]) > 1.5:
            continue
        if not (top <= edge["top"] <= bottom):
            continue
        width = edge["x1"] - edge["x0"]
        if width < 5 or width > plan_width * 0.98:
            continue
        if edge["x1"] < plan[0] - 40 or edge["x0"] > plan[1] + 40:
            continue
        low, high = _station_range(anchors, edge["x0"], edge["x1"])
        if low is not None:
            ranges.append((low, high))

    sheet.hca_reliable = True
    if not ranges:
        sheet.hca_ranges = []
        return

    ranges.sort()
    merged = [list(ranges[0])]
    for low, high in ranges[1:]:
        if low <= merged[-1][1] + 1:
            merged[-1][1] = max(merged[-1][1], high)
        else:
            merged.append([low, high])
    sheet.hca_ranges = [tuple(pair) for pair in merged]


def apply_alignment(dig, sheets) -> None:
    """Fill a dig's alignment-derived fields from the uploaded sheets.

    Two kinds of fact come off an alignment sheet and they are applied
    separately:

    * **Line-level** - the route name. It describes the pipeline, not a
      position on it, so it applies to every dig on that line and does not
      wait for a station match. Withholding it when no sheet happens to cover
      the dig's station was wrong.
    * **Station-level** - the sheet number, county band, PLSS band and HCA
      band. These do need the sheet that covers the dig.
    """
    if not sheets:
        dig.warnings.append(
            "No alignment sheets were uploaded, so line name, alignment "
            "sheet, tract, legal description and HCA are Unknown."
        )
        return

    _apply_line_level(dig, sheets)

    matches = [sheet for sheet in sheets if sheet.covers(dig.station_ft)]
    if not matches:
        unranged = [sheet for sheet in sheets if sheet.station_start is None]
        if len(sheets) == 1 and unranged:
            matches = list(sheets)
            dig.warnings.append(
                f"Could not read a station range from {sheets[0].sheet_id}, "
                "so it was used anyway. Check the tract, legal description "
                "and HCA."
            )
        else:
            dig.warnings.append(
                f"No alignment sheet covers station "
                f"{feet_to_station(dig.station_ft)}. Sheets read: "
                + "; ".join(sheet.describe() for sheet in sheets)
                + ". Alignment sheet, tract, legal description and HCA left "
                "as Unknown."
            )
            return

    sheet = matches[0]
    dig.alignment_sheet = sheet.sheet_id

    county = sheet.county_for(dig.station_ft)
    if county is not None:
        dig.county, dig.state = county.county, county.state
    else:
        if sheet.county != UNKNOWN:
            dig.county = sheet.county
        if sheet.state != UNKNOWN:
            dig.state = sheet.state

    segment = sheet.plss_for(dig.station_ft)
    if segment is not None:
        dig.legal_description = segment.label
        # On every filled report checked, Tract Number equals the PLSS section.
        dig.tract_number = segment.section
    elif len(sheet.plss) > 1:
        dig.warnings.append(
            "Alignment sheet spans more than one PLSS section and the station "
            "could not be placed; tract and legal description left as Unknown."
        )

    hca = sheet.hca_for(dig.station_ft)
    if hca == UNKNOWN and dig.hca_from_digsheet:
        hca = dig.hca_from_digsheet
    dig.hca = hca


def _apply_line_level(dig, sheets) -> None:
    """Route name, and county/state when every sheet agrees on them."""
    routes = []
    for sheet in sheets:
        if sheet.route_name != UNKNOWN and sheet.route_name not in routes:
            routes.append(sheet.route_name)
    if len(routes) == 1:
        dig.line_name = routes[0]
    elif len(routes) > 1:
        # Sheets from more than one line were uploaded; the covering sheet
        # decides, and if none covers this dig it stays Unknown.
        covering = [s for s in sheets if s.covers(dig.station_ft)]
        if covering and covering[0].route_name != UNKNOWN:
            dig.line_name = covering[0].route_name
        else:
            dig.warnings.append(
                "Alignment sheets from more than one line were uploaded ("
                + "; ".join(routes)
                + ") and none covers this dig, so the line name is Unknown."
            )

    counties = []
    for sheet in sheets:
        for entry in sheet.counties:
            pair = (entry.county, entry.state)
            if pair not in counties:
                counties.append(pair)
    if len(counties) == 1:
        dig.county, dig.state = counties[0]


# ==========================================================================
# KMZ
# ==========================================================================
# Read pipeline geometry out of a KMZ/KML export.
#
# Only two things are needed for the aerial image: the centerline itself and any
# point placemarks (AGMs, valves) close enough to the dig to be worth labelling.

KML_NS = {"kml": "http://www.opengis.net/kml/2.2"}


@dataclass
class PipelineData:
    """Centerlines and placemarks, all in (longitude, latitude) degrees."""

    lines: list = field(default_factory=list)      # list[list[tuple[float, float]]]
    placemarks: list = field(default_factory=list)  # list[tuple[str, float, float]]
    source_name: str = ""

    @property
    def is_empty(self) -> bool:
        return not self.lines and not self.placemarks


def _strip_namespace(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _parse_coordinates(text: str):
    points = []
    for chunk in re.split(r"\s+", (text or "").strip()):
        if not chunk:
            continue
        parts = chunk.split(",")
        if len(parts) < 2:
            continue
        try:
            longitude = float(parts[0])
            latitude = float(parts[1])
        except ValueError:
            continue
        points.append((longitude, latitude))
    return points


def parse_kml_bytes(data: bytes, source_name: str = "") -> PipelineData:
    result = PipelineData(source_name=source_name)
    root = ElementTree.fromstring(data)

    for element in root.iter():
        tag = _strip_namespace(element.tag)

        if tag in ("LineString", "LinearRing"):
            for child in element:
                if _strip_namespace(child.tag) == "coordinates":
                    points = _parse_coordinates(child.text)
                    if len(points) >= 2:
                        result.lines.append(points)

        elif tag == "Placemark":
            name = ""
            for child in element:
                if _strip_namespace(child.tag) == "name":
                    name = (child.text or "").strip()
                    break
            for descendant in element.iter():
                if _strip_namespace(descendant.tag) != "Point":
                    continue
                for child in descendant:
                    if _strip_namespace(child.tag) == "coordinates":
                        points = _parse_coordinates(child.text)
                        if points:
                            longitude, latitude = points[0]
                            result.placemarks.append((name, longitude, latitude))

    return result


def parse_kmz(data: bytes, filename: str = "") -> PipelineData:
    """Accepts a .kmz (zip) or a bare .kml."""
    if filename.lower().endswith(".kml") or data[:5] == b"<?xml":
        return parse_kml_bytes(data, filename)

    combined = PipelineData(source_name=filename)
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        names = [n for n in archive.namelist() if n.lower().endswith(".kml")]
        names.sort(key=lambda n: (n.lower() != "doc.kml", n))
        for name in names:
            try:
                part = parse_kml_bytes(archive.read(name), name)
            except ElementTree.ParseError:
                continue
            combined.lines.extend(part.lines)
            combined.placemarks.extend(part.placemarks)
    return combined


def find_placemark(data: PipelineData, name: str) -> Optional[tuple]:
    """A placemark by name, e.g. the AGM a dig references."""
    if not data or not name:
        return None
    wanted = re.sub(r"\s+", " ", name).strip().upper()
    for placemark_name, longitude, latitude in data.placemarks:
        if re.sub(r"\s+", " ", placemark_name or "").strip().upper() == wanted:
            return (placemark_name, longitude, latitude)
    return None


def nearest_placemark(data: PipelineData, latitude: float, longitude: float,
                      max_miles: float = 2.0, only_agms: bool = True) -> Optional[tuple]:
    """The closest placemark to a point, if one is near enough to label.

    A dig KMZ usually carries the digs themselves as placemarks alongside the
    AGMs, so by default only AGM-style names are considered - labelling the
    dig's own placemark as its AGM reference would be wrong.
    """
    if not data.placemarks:
        return None
    best, best_distance = None, None
    for name, plon, plat in data.placemarks:
        if only_agms and not AGM_NAME_RE.match((name or "").strip()):
            continue
        # Rough local-plane distance in miles; good enough for ranking.
        dx = (plon - longitude) * 54.6
        dy = (plat - latitude) * 69.0
        distance = (dx * dx + dy * dy) ** 0.5
        if best_distance is None or distance < best_distance:
            best, best_distance = (name, plon, plat), distance
    if best_distance is not None and best_distance <= max_miles:
        return best
    return None


# ==========================================================================
# AERIAL
# ==========================================================================
# Render the aerial image that goes into the staking report.
#
# Satellite tiles are composited, then the KMZ centerline, the nearest AGM and a
# labelled dig pin are drawn on top - the same information the Google Earth Pro
# screenshots carry today. The output is produced at the exact aspect ratio of
# the template's image slot so it drops in without distortion.

# The template's image slot: 5919108 x 2847624 EMU.
SLOT_ASPECT = 5919108 / 2847624

BASEMAPS = {
    # Mapbox uses the same token as the directions generator, and serves 512px
    # retina tiles, so it is the default where a token is configured.
    "Mapbox Satellite": {
        "url": "https://api.mapbox.com/v4/mapbox.satellite/{z}/{x}/{y}@2x.jpg90",
        "tile_size": 512,
        "needs_token": True,
    },
    "Esri World Imagery": {
        "url": (
            "https://server.arcgisonline.com/ArcGIS/rest/services/"
            "World_Imagery/MapServer/tile/{z}/{y}/{x}"
        ),
        "tile_size": 256,
        "needs_token": False,
    },
    "USDA NAIP (USA only)": {
        "url": (
            "https://gis.apfo.usda.gov/arcgis/rest/services/NAIP/"
            "USDA_CONUS_PRIME/ImageServer/tile/{z}/{y}/{x}"
        ),
        "tile_size": 256,
        "needs_token": False,
    },
}


def basemap_options(has_token: bool) -> list:
    return [
        name for name, spec in BASEMAPS.items()
        if has_token or not spec["needs_token"]
    ]

USER_AGENT = "ONEOK-Dig-File-Generator/1.0 (staking report automation)"

# Web Mercator maths is done in the tile set's own pixel units, so a retina
# (@2x, 512px) tile set simply doubles the resolution at a given zoom.
DEFAULT_TILE_SIZE = 256

PIPE_RED = (226, 32, 32)
LABEL_WHITE = (255, 255, 255)
SHADOW = (0, 0, 0)


@dataclass
class MapView:
    zoom: int
    centre_lat: float
    centre_lon: float
    width: int
    height: int
    tile_size: int = DEFAULT_TILE_SIZE

    def project(self, longitude: float, latitude: float) -> tuple[float, float]:
        """Longitude/latitude -> pixel coordinates in the rendered image."""
        cx, cy = _lonlat_to_world(self.centre_lon, self.centre_lat, self.zoom, self.tile_size)
        px, py = _lonlat_to_world(longitude, latitude, self.zoom, self.tile_size)
        return px - cx + self.width / 2.0, py - cy + self.height / 2.0


def _lonlat_to_world(longitude, latitude, zoom, tile_size=DEFAULT_TILE_SIZE):
    scale = tile_size * (2 ** zoom)
    x = (longitude + 180.0) / 360.0 * scale
    sin_lat = math.sin(math.radians(max(min(latitude, 85.05112878), -85.05112878)))
    y = (0.5 - math.log((1 + sin_lat) / (1 - sin_lat)) / (4 * math.pi)) * scale
    return x, y


def metres_per_pixel(latitude, zoom, tile_size=DEFAULT_TILE_SIZE) -> float:
    """Ground metres covered by one rendered pixel."""
    base = 156543.03392 * math.cos(math.radians(latitude)) / (2 ** zoom)
    return base * (DEFAULT_TILE_SIZE / tile_size)


def _fetch_tile(session, template: str, zoom: int, x: int, y: int, token: str = ""):
    url = template.format(z=zoom, x=x, y=y)
    if token:
        url += ("&" if "?" in url else "?") + f"access_token={token}"
    try:
        response = session.get(url, timeout=20, headers={"User-Agent": USER_AGENT})
        if response.status_code != 200 or not response.content:
            return None
        return Image.open(io.BytesIO(response.content)).convert("RGB")
    except Exception:
        return None


def _basemap_image(view: MapView, template: str, token: str = "") -> tuple:
    tile = view.tile_size
    cx, cy = _lonlat_to_world(view.centre_lon, view.centre_lat, view.zoom, tile)
    left = cx - view.width / 2.0
    top = cy - view.height / 2.0

    first_x = int(math.floor(left / tile))
    first_y = int(math.floor(top / tile))
    last_x = int(math.floor((left + view.width) / tile))
    last_y = int(math.floor((top + view.height) / tile))

    canvas = Image.new("RGB", (view.width, view.height), (32, 38, 32))
    jobs = [
        (x, y)
        for x in range(first_x, last_x + 1)
        for y in range(first_y, last_y + 1)
    ]
    if not jobs:
        return canvas, 0

    fetched = 0
    max_index = 2 ** view.zoom
    with requests.Session() as session:
        with ThreadPoolExecutor(max_workers=8) as pool:
            futures = {
                pool.submit(
                    _fetch_tile, session, template, view.zoom, x % max_index, y, token
                ): (x, y)
                for x, y in jobs
                if 0 <= y < max_index
            }
            for future, (x, y) in futures.items():
                image = future.result()
                if image is None:
                    continue
                if image.size != (tile, tile):
                    image = image.resize((tile, tile), Image.LANCZOS)
                canvas.paste(
                    image,
                    (int(round(x * tile - left)), int(round(y * tile - top))),
                )
                fetched += 1
    return canvas, fetched


# ---------------------------------------------------------------------------
# Drawing helpers
# ---------------------------------------------------------------------------

def _font(size: int):
    for path in (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/Arial.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ):
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _outlined_text(draw, xy, text, font, fill=LABEL_WHITE, outline=SHADOW, weight=2):
    x, y = xy
    for dx in range(-weight, weight + 1):
        for dy in range(-weight, weight + 1):
            if dx or dy:
                draw.text((x + dx, y + dy), text, font=font, fill=outline)
    draw.text((x, y), text, font=font, fill=fill)


def _centred_text(draw, centre, text, font, **kwargs):
    left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
    _outlined_text(
        draw,
        (centre[0] - (right - left) / 2.0, centre[1] - (bottom - top) / 2.0),
        text,
        font,
        **kwargs,
    )


def _draw_pin(draw, x, y, colour=(255, 214, 0)):
    """A small Google Earth style pushpin."""
    height = 22
    radius = 7
    draw.line([(x, y), (x, y - height + radius)], fill=(60, 60, 60), width=3)
    draw.ellipse(
        [x - radius, y - height, x + radius, y - height + 2 * radius],
        fill=colour,
        outline=(60, 60, 60),
        width=2,
    )


def _draw_flag(draw, x, y, colour=PIPE_RED):
    """A small flag marker, used for AGMs."""
    height = 24
    draw.line([(x, y), (x, y - height)], fill=(245, 245, 245), width=3)
    draw.polygon(
        [(x + 1, y - height), (x + 16, y - height + 6), (x + 1, y - height + 12)],
        fill=colour,
    )


def _scale_bar(draw, view, width, height):
    mpp = metres_per_pixel(view.centre_lat, view.zoom, view.tile_size)
    feet_per_pixel = mpp * 3.28084
    for target in (200, 400, 500, 800, 1000, 1500, 2000, 3000, 5000):
        pixels = target / feet_per_pixel
        if 90 <= pixels <= width * 0.22:
            break
    else:
        target, pixels = 400, 400 / feet_per_pixel

    right = width - 62
    left = right - pixels
    baseline = height - 18
    draw.line([(left, baseline), (right, baseline)], fill=LABEL_WHITE, width=3)
    draw.line([(left, baseline - 6), (left, baseline + 4)], fill=LABEL_WHITE, width=3)
    draw.line([(right, baseline - 6), (right, baseline + 4)], fill=LABEL_WHITE, width=3)
    font = _font(15)
    _outlined_text(draw, (left, baseline - 24), f"{target:,} ft", font, weight=1)


def _north_arrow(draw, width, height):
    font = _font(20)
    x = width - 28
    y = height - 62
    draw.polygon(
        [(x, y), (x - 8, y + 22), (x, y + 16), (x + 8, y + 22)],
        fill=LABEL_WHITE,
        outline=SHADOW,
    )
    _centred_text(draw, (x, y + 36), "N", font, weight=1)


def _legend(draw, entries, width):
    if not entries:
        return
    font = _font(15)
    title_font = _font(16)
    line_height = 20
    box_width = 190
    for label in entries:
        left, top, right, bottom = draw.textbbox((0, 0), label, font=font)
        box_width = max(box_width, int(right - left) + 46)
    box_height = 26 + line_height * len(entries)
    x0 = width - box_width - 14
    y0 = 12

    draw.rectangle(
        [x0, y0, x0 + box_width, y0 + box_height],
        fill=(255, 255, 255, 235),
        outline=(120, 120, 120),
    )
    draw.text((x0 + 10, y0 + 5), "Legend", font=title_font, fill=(20, 20, 20))
    for index, label in enumerate(entries):
        y = y0 + 26 + index * line_height
        draw.text((x0 + 30, y), label, font=font, fill=(20, 20, 20))
        draw.ellipse([x0 + 13, y + 5, x0 + 21, y + 13], fill=PIPE_RED)


def _title_card(draw, text):
    if not text:
        return
    font = _font(18)
    left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
    width = int(right - left) + 24
    height = int(bottom - top) + 20
    draw.rectangle([10, 10, 10 + width, 10 + height], fill=(255, 255, 255, 235),
                   outline=(120, 120, 120))
    draw.text((22, 18), text, font=font, fill=(20, 20, 20))


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

class AerialError(RuntimeError):
    pass


def render_aerial(
    dig,
    pipeline: Optional[PipelineData],
    width: int = 1600,
    basemap: str = "Esri World Imagery",
    span_feet: float = 2600.0,
    show_legend: bool = True,
    token: str = "",
) -> Optional[bytes]:
    """Return PNG bytes for one dig, or None when there is no coordinate."""
    if dig.ili_latitude is None or dig.ili_longitude is None:
        return None

    height = int(round(width / SLOT_ASPECT))
    latitude = float(dig.ili_latitude)
    longitude = float(dig.ili_longitude)

    spec = BASEMAPS.get(basemap) or BASEMAPS["Esri World Imagery"]
    tile_size = spec["tile_size"]
    tile_token = token if spec["needs_token"] else ""
    if spec["needs_token"] and not tile_token:
        raise AerialError(f"{basemap} needs a Mapbox token.")

    # Choose the zoom whose scale puts `span_feet` across the image width.
    target_mpp = (span_feet / 3.28084) / width
    zoom = 19
    for candidate in range(22, 9, -1):
        if metres_per_pixel(latitude, candidate, tile_size) >= target_mpp:
            zoom = candidate
            break

    view = MapView(zoom=zoom, centre_lat=latitude, centre_lon=longitude,
                   width=width, height=height, tile_size=tile_size)
    image, fetched = _basemap_image(view, spec["url"], tile_token)
    if fetched == 0:
        raise AerialError(
            f"No imagery tiles could be fetched from {basemap}. "
            "Check network access, or the Mapbox token if using Mapbox."
        )
    overlay = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay, "RGBA")

    legend_entries = []

    if pipeline is not None and pipeline.lines:
        drew_line = False
        for line in pipeline.lines:
            points = [view.project(lon, lat) for lon, lat in line]
            visible = [
                (x, y) for x, y in points
                if -width <= x <= width * 2 and -height <= y <= height * 2
            ]
            if len(visible) >= 2:
                draw.line(points, fill=PIPE_RED, width=4, joint="curve")
                drew_line = True
        if drew_line:
            legend_entries.append(_pipeline_legend_label(pipeline, dig))

    agm = None
    if pipeline is not None:
        # Prefer the AGM this dig actually references, then the nearest AGM.
        for reference in (dig.downstream_reference, dig.upstream_reference):
            agm = find_placemark(pipeline, reference)
            if agm is not None:
                break
        if agm is None:
            agm = nearest_placemark(pipeline, latitude, longitude)
    if agm is not None:
        name, alon, alat = agm
        ax, ay = view.project(alon, alat)
        if -50 <= ax <= width + 50 and -50 <= ay <= height + 50:
            _draw_flag(draw, ax, ay)
            _outlined_text(draw, (ax + 20, ay - 34), name or "AGM", _font(17))
            if name:
                legend_entries.append(name)
        else:
            agm = None

    dx, dy = view.project(longitude, latitude)
    _draw_pin(draw, dx, dy)
    _centred_text(draw, (dx, dy + 26), dig.name, _font(30))
    if dig.name:
        legend_entries.append(dig.name)

    if show_legend:
        _legend(draw, legend_entries, width)
    _title_card(draw, dig.name)
    _scale_bar(draw, view, width, height)
    _north_arrow(draw, width, height)

    image = Image.alpha_composite(image.convert("RGBA"), overlay).convert("RGB")

    buffer = io.BytesIO()
    image.save(buffer, format="PNG", optimize=True)
    return buffer.getvalue()


def _pipeline_legend_label(pipeline, dig) -> str:
    if dig.line_name and dig.line_name != "Unknown":
        head = dig.line_name.split(")")[0].lstrip("(")
        if head:
            return head.split("-")[0]
    name = (pipeline.source_name or "Pipeline").rsplit("/", 1)[-1]
    return name.rsplit(".", 1)[0] or "Pipeline"


# ==========================================================================
# DIRECTIONS
# ==========================================================================
# Driving directions, ported from Hayden's Dig Site Directions Generator.
#
# Source: github.com/maydenhiller/Dig-Site-Directions-Generator
#
# The logic is kept identical so the wording matches the reports already issued:
# reverse geocode the dig to its town, seed a route from the town centre to get a
# real road start point, name that start point by the two distinct roads meeting
# there, then phrase each step as "<instruction> and continue traveling
# <cardinal> for <n.nn> miles", closing with which side the dig sits on relative
# to the direction of travel.
#
# Needs a Mapbox token - the same one the existing app uses.

TIMEOUT = 25
CARDINALS = ["North", "Northeast", "East", "Southeast",
             "South", "Southwest", "West", "Northwest"]


class DirectionsError(RuntimeError):
    pass


@dataclass
class DirectionsResult:
    paragraph: str
    town: str = ""
    state: str = ""
    intersection: str = ""


# ---------------------------------------------------------------------------
# Web Mercator geometry - used to decide which side of the road the dig is on
# ---------------------------------------------------------------------------

def _mercator_xy(longitude: float, latitude: float) -> tuple:
    radius = 6378137.0
    x = math.radians(longitude) * radius
    y = math.log(math.tan(math.pi / 4 + math.radians(latitude) / 2)) * radius
    return x, y


def _segment_projection(a_lon, a_lat, b_lon, b_lat, p_lon, p_lat):
    ax, ay = _mercator_xy(a_lon, a_lat)
    bx, by = _mercator_xy(b_lon, b_lat)
    px, py = _mercator_xy(p_lon, p_lat)
    vx, vy = bx - ax, by - ay
    length_squared = vx * vx + vy * vy
    if length_squared == 0:
        return 0.0, ax, ay, ax, ay, bx, by
    t = ((px - ax) * vx + (py - ay) * vy) / length_squared
    t = max(0.0, min(1.0, t))
    return t, ax + t * vx, ay + t * vy, ax, ay, bx, by


def _nearest_segment(route, dig_lat, dig_lon):
    best, best_distance = None, float("inf")
    px, py = _mercator_xy(dig_lon, dig_lat)
    for index in range(len(route) - 1):
        a_lat, a_lon = route[index]
        b_lat, b_lon = route[index + 1]
        _, projx, projy, *_ = _segment_projection(
            a_lon, a_lat, b_lon, b_lat, dig_lon, dig_lat
        )
        dx, dy = px - projx, py - projy
        distance = dx * dx + dy * dy
        if distance < best_distance:
            best_distance = distance
            best = ((a_lon, a_lat), (b_lon, b_lat), (projx, projy))
    return best


def side_relative_to_route(route, dig_lat, dig_lon) -> str:
    nearest = _nearest_segment(route, dig_lat, dig_lon)
    if not nearest:
        return "right"
    (a_lon, a_lat), (b_lon, b_lat), (projx, projy) = nearest
    ax, ay = _mercator_xy(a_lon, a_lat)
    bx, by = _mercator_xy(b_lon, b_lat)
    px, py = _mercator_xy(dig_lon, dig_lat)
    cross = (bx - ax) * (py - projy) - (by - ay) * (px - projx)
    return "left" if cross > 0 else "right"


# ---------------------------------------------------------------------------
# Phrasing
# ---------------------------------------------------------------------------

def _bearing_to_cardinal(bearing: float) -> str:
    return CARDINALS[round((bearing % 360) / 45) % 8]


def _extract_town_state(feature) -> tuple:
    town, state = "", ""
    for context in feature.get("context", []):
        identifier = context.get("id", "")
        if identifier.startswith("place."):
            town = context.get("text", "")
        if identifier.startswith("region."):
            state = context.get("text", "")
    if not town:
        town = feature.get("text", "")
    return town, state


def _normalise_street_base(name: str) -> str:
    if not name:
        return ""
    text = re.sub(r"^(N|S|E|W)\s+", "", name.strip(), flags=re.I)
    text = re.sub(
        r"\b(Street|St\.?|Avenue|Ave\.?|Road|Rd\.?|Boulevard|Blvd\.?|Drive|Dr\.?"
        r"|Lane|Ln\.?|Terrace|Ter\.?|Court|Ct\.?)\b\.?",
        "",
        text,
        flags=re.I,
    ).strip()
    return re.sub(r"\s{2,}", " ", text)


def _intersection_label(longitude, latitude, token) -> str:
    url = (
        "https://api.mapbox.com/v4/mapbox.mapbox-streets-v8/tilequery/"
        f"{longitude},{latitude}.json?layers=road&radius=300&limit=50"
        f"&access_token={token}"
    )
    payload = requests.get(url, timeout=TIMEOUT).json()

    seen, roads = set(), []
    for feature in payload.get("features", []):
        properties = feature.get("properties", {})
        name = properties.get("name")
        if not name or name in seen:
            continue
        seen.add(name)
        roads.append((name, properties.get("class", ""), _normalise_street_base(name)))

    if not roads:
        return "Unknown Intersection"

    bases = [base for _, _, base in roads]
    if any("Washington" in b for b in bases) and any("Jefferson" in b for b in bases):
        washington = next(n for n, _, b in roads if "Washington" in b)
        jefferson = next(n for n, _, b in roads if "Jefferson" in b)
        return f"{washington} & {jefferson}"

    used, chosen = set(), []
    for name, _, base in roads:
        if base not in used:
            chosen.append(name)
            used.add(base)
        if len(chosen) == 2:
            break
    return " & ".join(chosen) if chosen else "Unknown Intersection"


def _format_step(step, miles: float) -> str:
    manoeuvre = step.get("maneuver", {})
    instruction = manoeuvre.get("instruction", "").rstrip(".")
    cardinal = _bearing_to_cardinal(manoeuvre.get("bearing_after", 0))
    return f"{instruction} and continue traveling {cardinal} for {miles:.2f} miles"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def generate_directions(latitude: float, longitude: float, token: str) -> DirectionsResult:
    if not token:
        raise DirectionsError("No Mapbox token configured.")

    import polyline

    town_url = (
        "https://api.mapbox.com/geocoding/v5/mapbox.places/"
        f"{longitude},{latitude}.json?types=place&language=en&access_token={token}"
    )
    town_payload = requests.get(town_url, timeout=TIMEOUT).json()
    features = town_payload.get("features") or []
    if not features:
        raise DirectionsError("Mapbox could not find a town near this coordinate.")
    town_feature = features[0]
    town, state = _extract_town_state(town_feature)
    town_centre = town_feature["center"]

    seed_url = (
        "https://api.mapbox.com/directions/v5/mapbox/driving/"
        f"{town_centre[0]},{town_centre[1]};{longitude},{latitude}"
        "?steps=true&geometries=polyline&overview=full&language=en"
        f"&access_token={token}"
    )
    seed = requests.get(seed_url, timeout=TIMEOUT).json()
    seed_routes = seed.get("routes") or []
    if not seed_routes:
        raise DirectionsError("Mapbox could not route to this coordinate.")
    start = seed_routes[0]["legs"][0]["steps"][0]["maneuver"]["location"]

    intersection = _intersection_label(start[0], start[1], token)

    route_url = (
        "https://api.mapbox.com/directions/v5/mapbox/driving/"
        f"{start[0]},{start[1]};{longitude},{latitude}"
        "?steps=true&geometries=polyline&overview=full&language=en"
        f"&access_token={token}"
    )
    payload = requests.get(route_url, timeout=TIMEOUT).json()
    routes = payload.get("routes") or []
    if not routes:
        raise DirectionsError("Mapbox returned no route from the start intersection.")
    route = routes[0]
    steps = route["legs"][0]["steps"]
    coordinates = polyline.decode(route["geometry"])

    narrative = [
        f"From the intersection of {intersection} in {town}, {state}, travel as follows"
    ]
    for index, step in enumerate(steps):
        if index == len(steps) - 1:
            side = side_relative_to_route(coordinates, latitude, longitude)
            narrative.append(f"The dig site will be located on your {side}.")
        else:
            narrative.append(_format_step(step, step["distance"] / 1609.34))

    paragraph = " ".join(line.strip().rstrip(".") + "." for line in narrative)
    return DirectionsResult(
        paragraph=paragraph, town=town, state=state, intersection=intersection
    )


# ==========================================================================
# CHEATSHEET
# ==========================================================================
# Fill the dig stake cheat sheet - three rows per dig, U/S weld, anomaly, D/S weld.
#
# The template ships in two flavours, ascending and descending station numbers.
# They differ only in the sign of the weld formulas, so rather than asking which
# one to upload, the correct formulas are written per dig from the stationing
# direction detected in that dig's own sheet. Either template can be used as the
# starting point.

FIRST_ROW = 6
ROWS_PER_DIG = 3

COL_NAME = 1        # A
COL_ODO = 2         # B
COL_STATION = 3     # C
COL_FROM_ANOMALY = 4  # D
COL_FROM_AGM = 5    # E
COL_NOTES = 6       # F
COL_HCA = 7         # G
COL_DEPTH = 8       # H
COL_LAT = 9         # I
COL_LON = 10        # J


def build_cheat_sheet(
    template_bytes: bytes,
    digs: Iterable[Dig],
    auto_notes: bool = True,
    proximity_feet: float = 2000.0,
) -> bytes:
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(template_bytes))
    sheet = workbook.worksheets[0]

    digs = list(digs)
    _ensure_rows(sheet, len(digs))

    previous: Optional[Dig] = None
    for index, dig in enumerate(digs):
        top = FIRST_ROW + index * ROWS_PER_DIG
        anomaly = top + 1
        bottom = top + 2

        sheet.cell(top, COL_NAME).value = f"{dig.name} U/S Weld"
        sheet.cell(anomaly, COL_NAME).value = f"{dig.name} Anomaly"
        sheet.cell(bottom, COL_NAME).value = f"{dig.name} D/S Weld"

        sheet.cell(anomaly, COL_ODO).value = dig.odometer
        sheet.cell(anomaly, COL_STATION).value = dig.station_ft

        # Weld distances sit on the weld rows; the formulas above and below
        # the anomaly derive the weld odometer and station from them.
        sheet.cell(top, COL_FROM_ANOMALY).value = dig.us_weld_distance
        sheet.cell(bottom, COL_FROM_ANOMALY).value = dig.ds_weld_distance

        # The upstream weld is always at a lower odometer, so the ODO formulas
        # never change. Only the station column flips, and that is the sole
        # difference between the ascending and descending templates.
        sheet.cell(top, COL_ODO).value = f"=B{anomaly}-D{top}"
        sheet.cell(bottom, COL_ODO).value = f"=B{anomaly}+D{bottom}"
        up, down = ("-", "+") if dig.station_ascends else ("+", "-")
        sheet.cell(top, COL_STATION).value = f"=C{anomaly}{up}D{top}"
        sheet.cell(bottom, COL_STATION).value = f"=C{anomaly}{down}D{bottom}"

        sheet.cell(anomaly, COL_FROM_AGM).value = dig.nearest_agm_label or None

        if auto_notes and previous is not None and dig.odometer and previous.odometer:
            gap = abs(dig.odometer - previous.odometer)
            if gap <= proximity_feet:
                sheet.cell(anomaly, COL_NOTES).value = (
                    f"{gap:.2f}' past {previous.name}"
                )
        if dig.notes:
            sheet.cell(anomaly, COL_NOTES).value = dig.notes

        sheet.cell(anomaly, COL_HCA).value = None if dig.hca == UNKNOWN else dig.hca

        # Latitude and longitude are field measurements. They are always left
        # blank for manual entry, and actively cleared so that a template that
        # already carries values in those columns cannot leak them through.
        for row in (top, anomaly, bottom):
            sheet.cell(row, COL_LAT).value = None
            sheet.cell(row, COL_LON).value = None

        previous = dig

    _clear_unused(sheet, len(digs))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _ensure_rows(sheet, dig_count: int) -> None:
    """Extend the template past its twelve pre-built dig blocks if needed."""
    needed = FIRST_ROW + dig_count * ROWS_PER_DIG - 1
    if needed <= sheet.max_row:
        return
    for row in range(sheet.max_row + 1, needed + 1):
        source_row = FIRST_ROW + ((row - FIRST_ROW) % ROWS_PER_DIG)
        for column in range(1, COL_LON + 1):
            source = sheet.cell(source_row, column)
            target = sheet.cell(row, column)
            if source.has_style:
                target._style = copy(source._style)


def _clear_unused(sheet, dig_count: int) -> None:
    """Blank out the template's leftover 'Dig #n' placeholder blocks."""
    start = FIRST_ROW + dig_count * ROWS_PER_DIG
    for row in range(start, sheet.max_row + 1):
        for column in range(1, COL_LON + 1):
            sheet.cell(row, column).value = None


# ==========================================================================
# STAKING
# ==========================================================================
# Write one filled staking report per dig, from the uploaded .xlsm template.

SHEET = "StakingReport"

DEFAULT_TOOL_TYPE = "Unknown Tool Type"
DEFAULT_STAKING_NOTES = (
    "The anomaly is marked with white paint and white flagging. The upstream "
    "and downstream welds are marked with pink paint and pink flagging."
)


@dataclass
class ReportSettings:
    surveyor_name: str = ""
    tool_type: str = DEFAULT_TOOL_TYPE
    staking_notes: str = DEFAULT_STAKING_NOTES
    write_line_name: bool = True
    write_county_state: bool = True


def build_staking_report(
    template_bytes: bytes,
    dig: Dig,
    settings: ReportSettings,
) -> tuple[bytes, list]:
    """Return (.xlsm bytes, warnings) for a single dig."""
    patcher = XlsmPatcher(template_bytes)
    warnings = []

    if SHEET not in patcher.sheet_names():
        raise ValueError(
            f"The template has no '{SHEET}' sheet - found "
            f"{', '.join(patcher.sheet_names())}."
        )

    values = {
        "G4": dig.name,
        "J3": settings.tool_type or DEFAULT_TOOL_TYPE,

        # Upstream / downstream references and the distance to each.
        "A5": dig.upstream_reference,
        "A6": dig.upstream_feet_to_agm,
        "N5": dig.downstream_reference,
        "N6": dig.downstream_feet_to_agm,

        # Odometer and stationing.
        "F5": dig.odometer,
        "K5": dig.station_ft,

        # Alignment sheet derived block.
        "C19": dig.alignment_sheet,
        "C20": _numeric_if_possible(dig.tract_number),
        "C21": dig.legal_description,
        "C24": dig.hca,

        # Free text.
        "A28": dig.directions or "",
        "J28": dig.staking_notes or settings.staking_notes,
    }

    if settings.write_line_name and dig.line_name and dig.line_name != UNKNOWN:
        values["J2"] = dig.line_name
    if settings.write_county_state:
        if dig.county and dig.county != UNKNOWN:
            values["C22"] = dig.county
        if dig.state and dig.state != UNKNOWN:
            values["C23"] = dig.state
    if settings.surveyor_name:
        # C11 only. The phone number in C12 is an XLOOKUP against the template's
        # own employee table, keyed off this name - writing C12 would replace
        # that formula with a static value.
        values["C11"] = settings.surveyor_name

    missing = patcher.set_values(SHEET, values)
    if missing:
        warnings.append(
            "Could not place these cells in the template: " + ", ".join(sorted(missing))
        )

    if dig.aerial_png:
        slot = patcher.find_image_slot(SHEET)
        if slot is None:
            warnings.append(
                "No image placeholder found on the template's StakingReport sheet, "
                "so the aerial image was not inserted."
            )
        else:
            patcher.replace_image(slot, _encode_for_slot(dig.aerial_png, slot))

    return patcher.to_bytes(), warnings


def _numeric_if_possible(value):
    if value is None:
        return None
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    return text


def _encode_for_slot(png_bytes: bytes, slot) -> bytes:
    """Match the placeholder's format and aspect so Excel does not stretch it."""

    image = Image.open(io.BytesIO(png_bytes)).convert("RGB")

    target_aspect = slot.aspect
    width, height = image.size
    if abs((width / height) - target_aspect) > 0.005:
        if width / height > target_aspect:
            new_width = int(round(height * target_aspect))
            left = (width - new_width) // 2
            image = image.crop((left, 0, left + new_width, height))
        else:
            new_height = int(round(width / target_aspect))
            top = (height - new_height) // 2
            image = image.crop((0, top, width, top + new_height))

    buffer = io.BytesIO()
    if slot.extension in ("jpg", "jpeg"):
        image.save(buffer, format="JPEG", quality=88, optimize=True)
    else:
        image.save(buffer, format="PNG", optimize=True)
    return buffer.getvalue()


# ==========================================================================
# STREAMLIT UI
# ==========================================================================

st.set_page_config(page_title="ONEOK Dig File Generator", page_icon="⛽", layout="wide")

STATE_KEYS = ["digs", "sheets", "pipeline", "outputs"]
for key in STATE_KEYS:
    st.session_state.setdefault(key, None)


def mapbox_token() -> str:
    try:
        return st.secrets["mapbox"]["token"]
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------

token = mapbox_token()

st.sidebar.header("Report details")
surveyor_name = st.sidebar.text_input(
    "Surveyed by",
    value="",
    help="The phone number fills itself in from the template's employee table.",
)
tool_type = st.sidebar.text_input("Tool type", value=DEFAULT_TOOL_TYPE)
county_entry = st.sidebar.text_input(
    "County",
    value="",
    help="Applied to every dig in the batch. Leave blank to use whatever the "
         "alignment sheet says.",
)
state_entry = st.sidebar.text_input(
    "State",
    value="",
    help="Applied to every dig in the batch. Leave blank to use whatever the "
         "alignment sheet says.",
)
staking_notes = st.sidebar.text_area("Staking notes", value=DEFAULT_STAKING_NOTES, height=110)

st.sidebar.header("Aerial image")
make_aerial = st.sidebar.checkbox("Generate aerial images", value=True)
basemap = st.sidebar.selectbox("Imagery", basemap_options(bool(token)), index=0)
span_feet = st.sidebar.slider("Image width across the ground (ft)", 800, 6000, 2600, 200)

st.sidebar.header("Directions")
make_directions = st.sidebar.checkbox(
    "Generate driving directions",
    value=bool(token),
    disabled=not token,
    help=(
        "Uses the same Mapbox logic as the Dig Site Directions Generator. "
        "Add a token to .streamlit/secrets.toml to enable."
    ),
)
if not token:
    st.sidebar.caption("No Mapbox token found — directions will be left blank.")

st.sidebar.header("Cheat sheet")
auto_notes = st.sidebar.checkbox("Auto-note digs close to the previous one", value=True)


# ---------------------------------------------------------------------------
# Uploads
# ---------------------------------------------------------------------------

st.title("ONEOK Dig File Generator")
st.caption(
    "Fills the staking report and cheat sheet from your dig sheets. "
    "Latitude, longitude, elevation, EDOC, survey date and photos are field "
    "measurements and are always left blank."
)

col_left, col_right = st.columns(2)
with col_left:
    template_file = st.file_uploader(
        "Staking report template (.xlsm)", type=["xlsm", "xlsx"], key="template"
    )
    cheat_file = st.file_uploader(
        "Cheat sheet template (.xlsx)", type=["xlsx"], key="cheat"
    )
    kmz_file = st.file_uploader(
        "Pipeline KMZ (for the aerial image)", type=["kmz", "kml"], key="kmz"
    )
with col_right:
    dig_files = st.file_uploader(
        "Dig sheets — PDF or Excel, several at once",
        type=["pdf", "xlsx", "xlsm", "xls"],
        accept_multiple_files=True,
        key="digs_upload",
    )
    alignment_files = st.file_uploader(
        "Alignment sheets (optional)",
        type=["pdf"],
        accept_multiple_files=True,
        key="alignment_upload",
    )

ready = bool(template_file and cheat_file and dig_files)
if not ready:
    st.info(
        "Upload the staking report template, the cheat sheet template and at "
        "least one dig sheet to begin."
    )


# ---------------------------------------------------------------------------
# Read the uploads
# ---------------------------------------------------------------------------

if st.button("Read uploads", type="primary", disabled=not ready):
    digs = []
    problems = []
    diagnostics = []

    with st.status("Reading dig sheets…", expanded=False):
        for upload in dig_files:
            try:
                found = parse_digsheet(upload.getvalue(), upload.name)
            except Exception as error:  # noqa: BLE001 - surfaced to the user
                problems.append(f"{upload.name}: {error}")
                continue
            if not found:
                if upload.name.lower().endswith(".pdf"):
                    report = pdf_digsheet_report(upload.getvalue(), upload.name)
                    diagnostics.append(report)
                    problems.append(
                        f"{upload.name}: no anomaly row — "
                        + report.get("why_nothing_matched", "reason unknown.")
                    )
                else:
                    problems.append(
                        f"{upload.name}: no row carrying a Dig Number was found."
                    )
            digs.extend(found)

    sheets = []
    if alignment_files:
        with st.status("Reading alignment sheets…", expanded=False):
            for upload in alignment_files:
                try:
                    sheets.append(parse_alignment_sheet(upload.getvalue(), upload.name))
                except Exception as error:  # noqa: BLE001
                    problems.append(f"{upload.name}: {error}")

    for dig in digs:
        apply_alignment(dig, sheets)
        dig.staking_notes = staking_notes
        # Typed values win over anything read off an alignment sheet.
        if county_entry.strip():
            dig.county = county_entry.strip()
        if state_entry.strip():
            dig.state = state_entry.strip()

    pipeline = None
    if kmz_file is not None:
        try:
            pipeline = parse_kmz(kmz_file.getvalue(), kmz_file.name)
            if pipeline.is_empty:
                problems.append(f"{kmz_file.name}: no lines or placemarks found.")
                pipeline = None
        except Exception as error:  # noqa: BLE001
            problems.append(f"{kmz_file.name}: {error}")

    digs.sort(key=lambda d: (d.source_file, d.odometer or 0))
    st.session_state.digs = digs
    st.session_state.sheets = sheets
    st.session_state.pipeline = pipeline
    st.session_state.outputs = None

    for problem in problems:
        st.warning(problem)

    if diagnostics:
        with st.expander("What the parser saw in those PDFs", expanded=True):
            st.caption(
                "Send this to Claude and the parser can be corrected against "
                "your actual sheet without guessing."
            )
            for report in diagnostics:
                st.json(report)


digs = st.session_state.digs or []

if digs:
    st.subheader(f"{len(digs)} dig{'s' if len(digs) != 1 else ''} found")

    sheets = st.session_state.sheets or []
    if sheets:
        with st.expander("Alignment sheets read", expanded=False):
            for sheet in sheets:
                station_range = (
                    f"{sheet.station_start}–{sheet.station_end}"
                    if sheet.station_start is not None
                    else "station range not found"
                )
                st.write(
                    f"**{sheet.sheet_id}** · {sheet.route_name} · "
                    f"{sheet.county}, {sheet.state} · {station_range} · "
                    f"HCA band {'read' if sheet.hca_reliable else 'not readable'}"
                )

    st.caption(
        "Check every value before generating. Fields the uploads could not "
        "establish are marked Unknown."
    )

    edited = st.data_editor(
        [
            {
                "Dig": dig.name,
                "ODO": dig.odometer,
                "Station": dig.station_ft,
                "U/S ref": dig.upstream_reference,
                "U/S ft": dig.upstream_feet_to_agm,
                "D/S ref": dig.downstream_reference,
                "D/S ft": dig.downstream_feet_to_agm,
                "Line name": dig.line_name,
                "Alignment sheet": dig.alignment_sheet,
                "Tract": dig.tract_number,
                "Legal description": dig.legal_description,
                "County": dig.county,
                "State": dig.state,
                "HCA": dig.hca,
                "Notes": dig.notes,
            }
            for dig in digs
        ],
        width="stretch",
        num_rows="fixed",
        key="review",
    )

    warnings = [(dig.name, warning) for dig in digs for warning in dig.warnings]
    if warnings:
        with st.expander(f"{len(warnings)} thing(s) to check", expanded=True):
            for name, warning in warnings:
                st.write(f"**{name}** — {warning}")

    # ---------------------------------------------------------------
    # Generate
    # ---------------------------------------------------------------
    if st.button("Generate files", type="primary"):
        for row, dig in zip(edited, digs):
            dig.line_name = row["Line name"]
            dig.alignment_sheet = row["Alignment sheet"]
            dig.tract_number = row["Tract"]
            dig.legal_description = row["Legal description"]
            dig.county = row["County"]
            dig.state = row["State"]
            dig.hca = row["HCA"]
            dig.notes = row["Notes"] or ""

        settings = ReportSettings(
            surveyor_name=surveyor_name,
            tool_type=tool_type,
            staking_notes=staking_notes,
        )

        template_bytes = template_file.getvalue()
        cheat_bytes = cheat_file.getvalue()
        pipeline = st.session_state.pipeline
        outputs = {}
        issues = []

        progress = st.progress(0.0, text="Starting…")
        total = len(digs)

        for index, dig in enumerate(digs, start=1):
            progress.progress((index - 0.5) / total, text=f"{dig.name} — aerial image")

            if make_aerial:
                try:
                    dig.aerial_png = render_aerial(
                        dig,
                        pipeline,
                        basemap=basemap,
                        span_feet=float(span_feet),
                        token=token,
                    )
                    if dig.aerial_png is None and dig.ili_latitude is not None:
                        issues.append(f"{dig.name}: aerial image could not be rendered.")
                except Exception as error:  # noqa: BLE001
                    issues.append(f"{dig.name}: aerial image failed — {error}")

            if make_directions and token and dig.ili_latitude is not None:
                progress.progress((index - 0.25) / total, text=f"{dig.name} — directions")
                try:
                    result = generate_directions(
                        float(dig.ili_latitude), float(dig.ili_longitude), token
                    )
                    dig.directions = result.paragraph
                except Exception as error:  # noqa: BLE001
                    issues.append(f"{dig.name}: directions failed — {error}")

            progress.progress(index / total, text=f"{dig.name} — staking report")
            try:
                report, report_warnings = build_staking_report(
                    template_bytes, dig, settings
                )
                outputs[f"{dig.output_basename}.xlsm"] = report
                issues.extend(f"{dig.name}: {w}" for w in report_warnings)
            except Exception as error:  # noqa: BLE001
                issues.append(f"{dig.name}: staking report failed — {error}")

        try:
            outputs["Dig Stake Cheat Sheet.xlsx"] = build_cheat_sheet(
                cheat_bytes,
                digs,
                auto_notes=auto_notes,
            )
        except Exception as error:  # noqa: BLE001
            issues.append(f"Cheat sheet failed — {error}")

        progress.empty()
        st.session_state.outputs = outputs
        for issue in issues:
            st.warning(issue)


# ---------------------------------------------------------------------------
# Downloads
# ---------------------------------------------------------------------------

outputs = st.session_state.outputs or {}
if outputs:
    st.subheader("Files")

    bundle = io.BytesIO()
    with zipfile.ZipFile(bundle, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in outputs.items():
            archive.writestr(name, data)

    st.download_button(
        "Download everything (.zip)",
        data=bundle.getvalue(),
        file_name=f"Dig Files {date.today():%Y-%m-%d}.zip",
        mime="application/zip",
        type="primary",
    )

    for name, data in outputs.items():
        mime = (
            "application/vnd.ms-excel.sheet.macroEnabled.12"
            if name.endswith(".xlsm")
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.download_button(name, data=data, file_name=name, mime=mime, key=f"dl_{name}")

    previews = [dig for dig in (st.session_state.digs or []) if dig.aerial_png]
    if previews:
        with st.expander("Aerial images", expanded=False):
            for dig in previews:
                st.image(dig.aerial_png, caption=dig.name, width="stretch")
